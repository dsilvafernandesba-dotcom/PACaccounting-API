from typing import Any, Dict, List, Optional, Set

import requests
from fastapi import APIRouter, Form, Request, UploadFile, File, HTTPException, Body
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from io import BytesIO
from openpyxl import load_workbook

from dados import estado, guardar_dados

router = APIRouter()
templates = Jinja2Templates(directory="templates")

# ====== API KEY do NIF.pt ======
API_KEY_NIF = "6467ab6c27daea9cb1219aca828c748b"


# ====== Filtro para formatar em euros (1.234,56 €) ======

def format_eur(valor: Any) -> str:
    try:
        n = float(valor)
    except Exception:
        n = 0.0
    s = f"{n:,.2f}"  # 1,234.56
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return s + " €"


templates.env.filters["eur"] = format_eur

# ====== Listas por defeito (dropdowns) ======

DEFAULT_LISTAS = {
    "carteiras": ["Pedro Fernandes", "Ana Rodrigues"],
    "tecnicos": ["Pedro Fernandes", "Ana Rodrigues"],
    "tecn_grh": [],
    "tipos_contabilidade": ["Organizada", "Simplificada"],
    "periodicidades_iva": ["Mensal", "Trimestral", "Anual"],
    "regimes_iva": ["Regime Normal", "Isento art. 53.º", "Outros"],
}


def obter_listas_opcoes() -> Dict[str, List[str]]:
    listas = estado.setdefault("listas", {})
    for chave, valores in DEFAULT_LISTAS.items():
        if chave not in listas or not isinstance(listas[chave], list):
            listas[chave] = list(valores)
    return listas


def normalizar_cliente(c: Dict[str, Any] | None) -> Dict[str, Any]:
    base: Dict[str, Any] = {
        "nome": "",
        "nif": "",
        "mensalidade": 0.0,
        "valor_grh": 0.0,
        "valor_toconline": 0.0,  # usamos como "G. Comercial"
        "carteira": "",
        "tecnico": "",
        "tecnico_grh": "",
        "tipo_contabilidade": "",
        "periodicidade_iva": "",
        "regime_iva": "",
        "com_fatura": True,
    }
    if c:
        base.update(c)
    return base


def normalizar_nome(nome: str) -> str:
    """
    Se o nome estiver TODO em maiúsculas, converte para Title Case.
    Caso contrário, deixa como está (tirando espaços a mais).
    """
    nome = (nome or "").strip()
    if not nome:
        return nome
    if nome == nome.upper():
        return nome.title()
    return nome


def _cliente_chave(cliente: Dict[str, Any]) -> Optional[str]:
    """Gera uma chave estável (preferindo NIF) para identificar clientes."""
    nif = str(cliente.get("nif") or "").strip()
    if nif:
        return f"NIF::{nif}"

    nome = str(cliente.get("nome") or "").strip()
    if nome:
        return f"NOME::{nome.upper()}"

    return None


def _merge_dados_cliente(destino: Dict[str, Any], origem: Dict[str, Any]) -> None:
    """Atualiza campos do cliente mantendo chaves existentes não mencionadas."""
    for chave, valor in origem.items():
        if chave == "_idx":
            continue
        destino[chave] = valor


# ================== LISTA DE CLIENTES (com filtros + totais) ==================

@router.get("/clientes", response_class=HTMLResponse)
async def pagina_clientes(
    request: Request,
    carteira: str = "",
    tecnico: str = "",
    tipo_contabilidade: str = "",
    periodicidade_iva: str = "",
    regime_iva: str = "",
    com_fatura: str = "",
):
    listas = obter_listas_opcoes()

    # Obtemos a lista original com índices reais
    clientes_orig = estado.get("clientes", [])
    enumerados = list(enumerate(clientes_orig))  # [(idx_global, dict_cliente), ...]

    # Ordenamos por nome, mas preservando o índice global
    ordenados = sorted(
        enumerados,
        key=lambda par: (par[1].get("nome") or "").upper()
    )

    filtrados: List[Dict[str, Any]] = []

    total_mensalidade = 0.0
    total_grh = 0.0
    total_gcomercial = 0.0

    for idx_global, c in ordenados:
        # Normalizar nome para nunca ficar em CAPS LOCK
        nome_atual = c.get("nome") or ""
        nome_norm = normalizar_nome(nome_atual)
        if nome_norm != nome_atual:
            c["nome"] = nome_norm  # atualiza em memória (e fica normalizado na UI)

        # Aplicar filtros
        if carteira and (c.get("carteira") or "") != carteira:
            continue
        if tecnico and (c.get("tecnico") or "") != tecnico:
            continue
        if tipo_contabilidade and (c.get("tipo_contabilidade") or "") != tipo_contabilidade:
            continue
        if periodicidade_iva and (c.get("periodicidade_iva") or "") != periodicidade_iva:
            continue
        if regime_iva and (c.get("regime_iva") or "") != regime_iva:
            continue
        if com_fatura:
            cf = bool(c.get("com_fatura"))
            if com_fatura == "sim" and not cf:
                continue
            if com_fatura == "nao" and cf:
                continue

        # Guardamos o índice real dentro do próprio dict
        c["_idx"] = idx_global
        filtrados.append(c)

        total_mensalidade += float(c.get("mensalidade") or 0.0)
        total_grh += float(c.get("valor_grh") or 0.0)
        total_gcomercial += float(c.get("valor_toconline") or 0.0)

    filtros = {
        "carteira": carteira,
        "tecnico": tecnico,
        "tipo_contabilidade": tipo_contabilidade,
        "periodicidade_iva": periodicidade_iva,
        "regime_iva": regime_iva,
        "com_fatura": com_fatura,
    }

    totais = {
        "mensalidade": total_mensalidade,
        "valor_grh": total_grh,
        "valor_gcomercial": total_gcomercial,
    }

    return templates.TemplateResponse(
        "clientes.html",
        {
            "request": request,
            "clientes": filtrados,
            "listas": listas,
            "filtros": filtros,
            "totais": totais,
        },
    )


# ================== NOVO CLIENTE ==================

@router.get("/clientes/novo", response_class=HTMLResponse)
async def pagina_novo_cliente(request: Request):
    listas = obter_listas_opcoes()
    cliente = normalizar_cliente(None)
    return templates.TemplateResponse(
        "cliente_form.html",
        {
            "request": request,
            "cliente": cliente,
            "idx": None,
            "listas": listas,
            "modo": "novo",
        },
    )


@router.post("/clientes/adicionar")
async def adicionar_cliente(
    nome: str = Form(...),
    nif: str = Form(...),
    mensalidade: float = Form(0.0),
    valor_grh: float = Form(0.0),
    valor_toconline: float = Form(0.0),  # G. Comercial
    carteira: str = Form(""),
    tecnico: str = Form(""),
    tecnico_grh: str = Form(""),
    tipo_contabilidade: str = Form(""),
    periodicidade_iva: str = Form(""),
    regime_iva: str = Form(""),
    com_fatura: str = Form("sim"),
):
    lista = estado.setdefault("clientes", [])
    nome_fmt = normalizar_nome(nome)
    nif_limpo = nif.strip()

    lista.append(
        {
            "nome": nome_fmt,
            "nif": nif_limpo,
            "mensalidade": float(mensalidade),
            "valor_grh": float(valor_grh),
            "valor_toconline": float(valor_toconline),
            "carteira": carteira.strip() or None,
            "tecnico": tecnico.strip() or None,
            "tecnico_grh": tecnico_grh.strip() or None,
            "tipo_contabilidade": tipo_contabilidade.strip() or None,
            "periodicidade_iva": periodicidade_iva.strip() or None,
            "regime_iva": regime_iva.strip() or None,
            "com_fatura": (com_fatura == "sim"),
        }
    )
    guardar_dados()
    # Depois de gravar, volta à lista e faz scroll ao cliente pelo NIF
    return RedirectResponse(url=f"/clientes#cliente-{nif_limpo}", status_code=303)


# ================== EDITAR CLIENTE ==================

@router.get("/clientes/editar/{idx}", response_class=HTMLResponse)
async def pagina_editar_cliente(request: Request, idx: int):
    clientes = estado.get("clientes", [])
    if not (0 <= idx < len(clientes)):
        return RedirectResponse(url="/clientes", status_code=303)
    listas = obter_listas_opcoes()
    cliente = normalizar_cliente(clientes[idx])
    return templates.TemplateResponse(
        "cliente_form.html",
        {
            "request": request,
            "cliente": cliente,
            "idx": idx,
            "listas": listas,
            "modo": "editar",
        },
    )


@router.post("/clientes/atualizar")
async def atualizar_cliente(
    idx: int = Form(...),
    nome: str = Form(...),
    nif: str = Form(...),
    mensalidade: float = Form(0.0),
    valor_grh: float = Form(0.0),
    valor_toconline: float = Form(0.0),  # G. Comercial
    carteira: str = Form(""),
    tecnico: str = Form(""),
    tecnico_grh: str = Form(""),
    tipo_contabilidade: str = Form(""),
    periodicidade_iva: str = Form(""),
    regime_iva: str = Form(""),
    com_fatura: str = Form("sim"),
):
    clientes = estado.get("clientes", [])
    if 0 <= idx < len(clientes):
        nome_fmt = normalizar_nome(nome)
        nif_limpo = nif.strip()
        clientes[idx] = {
            "nome": nome_fmt,
            "nif": nif_limpo,
            "mensalidade": float(mensalidade),
            "valor_grh": float(valor_grh),
            "valor_toconline": float(valor_toconline),
            "carteira": carteira.strip() or None,
            "tecnico": tecnico.strip() or None,
            "tecnico_grh": tecnico_grh.strip() or None,
            "tipo_contabilidade": tipo_contabilidade.strip() or None,
            "periodicidade_iva": periodicidade_iva.strip() or None,
            "regime_iva": regime_iva.strip() or None,
            "com_fatura": (com_fatura == "sim"),
        }
        guardar_dados()
        # Depois de gravar, volta à lista e faz scroll ao cliente pelo NIF
        return RedirectResponse(url=f"/clientes#cliente-{nif_limpo}", status_code=303)

    return RedirectResponse(url="/clientes", status_code=303)


# ================== IMPORTAR CLIENTES POR EXCEL ==================

@router.post("/clientes/importar-excel")
async def importar_clientes_excel(ficheiro: UploadFile = File(...)):
    """
    Importa clientes a partir de um ficheiro Excel com colunas:
    nome, nif, mensalidade, valor_grh, valor_gestao_comercial,
    carteira, tecnico, tipo_contabilidade, periodicidade_iva, com_fatura
    (modelo v2 que combinámos).
    """
    if not ficheiro.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Ficheiro inválido. Use um Excel (.xlsx/.xls).")

    dados_bytes = await ficheiro.read()
    wb = load_workbook(filename=BytesIO(dados_bytes), data_only=True)
    ws = wb.active

    # Cabeçalhos na 1.ª linha
    header = []
    for cell in ws[1]:
        header.append(str(cell.value).strip().lower() if cell.value is not None else "")

    def idx(nome_coluna: str):
        nome_coluna = nome_coluna.lower()
        return header.index(nome_coluna) if nome_coluna in header else None

    idx_nome = idx("nome")
    idx_nif = idx("nif")
    idx_mensalidade = idx("mensalidade")
    idx_valor_grh = idx("valor_grh")
    idx_valor_gc = idx("valor_gestao_comercial")
    idx_carteira = idx("carteira")
    idx_tecnico = idx("tecnico")
    idx_tipo_cont = idx("tipo_contabilidade")
    idx_per_iva = idx("periodicidade_iva")
    idx_com_fat = idx("com_fatura")

    dados_local = estado if isinstance(estado, dict) else {}
    lista = dados_local.setdefault("clientes", [])

    def to_float(v):
        try:
            return float(v)
        except Exception:
            return 0.0

    def parse_bool(v) -> bool:
        if v is None:
            return True
        s = str(v).strip().lower()
        if s in ("1", "true", "verdadeiro", "sim", "s", "y", "yes"):
            return True
        if s in ("0", "false", "falso", "nao", "não", "n", "no"):
            return False
        return True

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        nome_val = row[idx_nome] if idx_nome is not None else ""
        nif_val = row[idx_nif] if idx_nif is not None else ""

        nome = normalizar_nome(str(nome_val or ""))
        nif = str(nif_val or "").strip()

        if not nome and not nif:
            continue

        mensalidade = to_float(row[idx_mensalidade]) if idx_mensalidade is not None else 0.0
        valor_grh = to_float(row[idx_valor_grh]) if idx_valor_grh is not None else 0.0
        valor_gc = to_float(row[idx_valor_gc]) if idx_valor_gc is not None else 0.0
        carteira = str(row[idx_carteira]).strip() if idx_carteira is not None and row[idx_carteira] is not None else ""
        tecnico = str(row[idx_tecnico]).strip() if idx_tecnico is not None and row[idx_tecnico] is not None else ""
        tipo_cont = str(row[idx_tipo_cont]).strip() if idx_tipo_cont is not None and row[idx_tipo_cont] is not None else ""
        per_iva = str(row[idx_per_iva]).strip() if idx_per_iva is not None and row[idx_per_iva] is not None else ""
        com_fat_val = row[idx_com_fat] if idx_com_fat is not None else True

        cliente_novo = {
            "nome": nome,
            "nif": nif,
            "mensalidade": mensalidade,
            "valor_grh": valor_grh,
            "valor_toconline": valor_gc,  # G. Comercial alimentado a partir do Excel
            "carteira": carteira or None,
            "tecnico": tecnico or None,
            "tecnico_grh": None,
            "tipo_contabilidade": tipo_cont or None,
            "periodicidade_iva": per_iva or None,
            "regime_iva": None,  # não vem no Excel, fica em branco
            "com_fatura": parse_bool(com_fat_val),
        }

        # Se já existir cliente com o mesmo NIF, atualiza; senão adiciona
        atualizado = False
        if nif:
            for i, c in enumerate(lista):
                if (c.get("nif") or "").strip() == nif:
                    lista[i] = cliente_novo
                    atualizado = True
                    break

        if not atualizado:
            lista.append(cliente_novo)

    guardar_dados()
    return RedirectResponse(url="/clientes", status_code=303)


@router.post("/clientes/sincronizar")
async def sincronizar_clientes_api(payload: Any = Body(...)):
    """Sincroniza lista de clientes preservando todo o estado restante."""
    full_sync = False
    remover_nifs_input: Set[str] = set()

    if isinstance(payload, list):
        recebidos = payload
    elif isinstance(payload, dict):
        recebidos = payload.get("clientes")
        full_sync = bool(payload.get("full_sync") or payload.get("replace"))
        remover_raw = payload.get("remover_nifs")
        if isinstance(remover_raw, (list, tuple, set)):
            remover_nifs_input = {
                str(nif).strip()
                for nif in remover_raw
                if str(nif).strip()
            }
    else:
        recebidos = None

    if not isinstance(recebidos, list):
        raise HTTPException(status_code=400, detail="Corpo inválido: esperado array de clientes.")

    clientes_antes = estado.get("clientes", [])
    if not isinstance(clientes_antes, list):
        clientes_antes = []

    estado_chaves_antes = sorted(list(estado.keys()))
    print(
        "[CLIENTES][SYNC][DEBUG] antes",
        {
            "total_clientes": len(clientes_antes),
            "keys": estado_chaves_antes,
        },
    )

    mapa_por_nif: Dict[str, Dict[str, Any]] = {}
    mapa_original_por_nif: Dict[str, Dict[str, Any]] = {}
    nifs_antes: Set[str] = set()
    clientes_sem_nif: List[Dict[str, Any]] = []
    info_sync_por_nif: Dict[str, Dict[str, Optional[str]]] = {}

    for cli in clientes_antes:
        nif = str(cli.get("nif") or "").strip()
        cli_sem_idx = dict(cli)
        cli_sem_idx.pop("_idx", None)
        if nif:
            mapa_por_nif[nif] = cli_sem_idx
            mapa_original_por_nif[nif] = dict(cli_sem_idx)
            nifs_antes.add(nif)
        else:
            clientes_sem_nif.append(cli_sem_idx)

    novos_clientes: List[Dict[str, Any]] = []
    nifs_depois: Set[str] = set()
    ordem_recebidos: List[str] = []

    for item in recebidos:
        if not isinstance(item, dict):
            continue

        nif = str(item.get("nif") or "").strip()
        if not nif or nif in nifs_depois:
            continue

        nifs_depois.add(nif)
        ordem_recebidos.append(nif)
        base_existente = mapa_por_nif.get(nif)
        cliente_base = normalizar_cliente(base_existente)
        if base_existente:
            _merge_dados_cliente(cliente_base, base_existente)
        _merge_dados_cliente(cliente_base, item)
        cliente_base["nome"] = normalizar_nome(cliente_base.get("nome", ""))
        cliente_base.pop("_idx", None)
        mapa_por_nif[nif] = cliente_base
        nome_antigo = None
        if base_existente:
            nome_antigo = str(base_existente.get("nome") or "").strip() or None
        nome_novo = str(cliente_base.get("nome") or "").strip() or None
        info_sync_por_nif[nif] = {
            "nome_antigo": nome_antigo,
            "nome_novo": nome_novo,
        }
        novos_clientes.append(cliente_base)

    remover_nifs = set(remover_nifs_input)

    if full_sync:
        # Lista final segue a ordem recebida, removendo apenas os NIFs excluídos, mantendo clientes sem NIF.
        clientes_resultantes: List[Dict[str, Any]] = []
        usados_nifs: Set[str] = set()
        for nif in ordem_recebidos:
            if nif in remover_nifs:
                continue
            cliente_atual = mapa_por_nif.get(nif)
            if cliente_atual:
                clientes_resultantes.append(cliente_atual)
                usados_nifs.add(nif)
        for cli in clientes_sem_nif:
            clientes_resultantes.append(cli)
    else:
        # Incremental: mantém ordem existente, atualizando/adicionando apenas o que chegou.
        clientes_resultantes = []
        usados_nifs = set()
        for cli in clientes_antes:
            nif = str(cli.get("nif") or "").strip()
            cli_sem_idx = dict(cli)
            cli_sem_idx.pop("_idx", None)
            if nif:
                if nif in remover_nifs:
                    continue
                atualizado = mapa_por_nif.get(nif)
                if atualizado:
                    clientes_resultantes.append(atualizado)
                else:
                    clientes_resultantes.append(cli_sem_idx)
                usados_nifs.add(nif)
            else:
                clientes_resultantes.append(cli_sem_idx)

        for nif in ordem_recebidos:
            if nif in usados_nifs or nif in remover_nifs:
                continue
            cliente_atual = mapa_por_nif.get(nif)
            if cliente_atual:
                clientes_resultantes.append(cliente_atual)
                usados_nifs.add(nif)

    estado["clientes"] = clientes_resultantes

    timings_dados = estado.get("timings_dados")
    if isinstance(timings_dados, dict):
        # Mantém timings coerentes com os clientes ativos e renomeados.
        for nif, info in info_sync_por_nif.items():
            nome_novo = info.get("nome_novo") or ""
            if not nome_novo:
                continue
            nome_antigo = info.get("nome_antigo") or ""
            for ano_dict in timings_dados.values():
                if not isinstance(ano_dict, dict):
                    continue
                reg_ativo = None
                if nome_antigo and nome_antigo != nome_novo and nome_antigo in ano_dict:
                    reg_antigo = ano_dict.pop(nome_antigo)
                    if isinstance(reg_antigo, dict):
                        existente = ano_dict.get(nome_novo)
                        if isinstance(existente, dict):
                            meses_antigos = reg_antigo.get("meses")
                            if isinstance(meses_antigos, dict):
                                destino_meses = existente.setdefault("meses", {})
                                if isinstance(destino_meses, dict):
                                    destino_meses.update(meses_antigos)
                            por_tecnico_antigo = reg_antigo.get("por_tecnico")
                            if isinstance(por_tecnico_antigo, dict):
                                destino_por = existente.setdefault("por_tecnico", {})
                                if isinstance(destino_por, dict):
                                    for tecnico, meses in por_tecnico_antigo.items():
                                        if isinstance(meses, dict):
                                            destino_tecnico = destino_por.setdefault(tecnico, {})
                                            if isinstance(destino_tecnico, dict):
                                                destino_tecnico.update(meses)
                                            else:
                                                destino_por[tecnico] = meses
                                        else:
                                            destino_por[tecnico] = meses
                        else:
                            ano_dict[nome_novo] = reg_antigo
                            existente = reg_antigo
                        reg_ativo = ano_dict.get(nome_novo)
                    else:
                        ano_dict[nome_antigo] = reg_antigo
                if reg_ativo is None:
                    reg_ativo = ano_dict.get(nome_novo)
                if isinstance(reg_ativo, dict):
                    reg_ativo.pop("apagado", None)

    nifs_finais = {
        str(c.get("nif") or "").strip()
        for c in clientes_resultantes
        if str(c.get("nif") or "").strip()
    }

    removidos_nifs: Set[str] = set()
    if full_sync:
        removidos_nifs.update(nifs_antes - nifs_finais)
    if remover_nifs:
        removidos_nifs.update(remover_nifs & nifs_antes)

    if removidos_nifs:
        if isinstance(timings_dados, dict):
            for nif in removidos_nifs:
                antigo = mapa_original_por_nif.get(nif) or mapa_por_nif.get(nif)
                nome_antigo = antigo.get("nome") if antigo else None
                if not nome_antigo:
                    continue
                for ano_dict in timings_dados.values():
                    if not isinstance(ano_dict, dict):
                        continue
                    # Marca como apagado comparando com versões normalizadas do nome.
                    for chave_nome, reg in ano_dict.items():
                        if not isinstance(reg, dict):
                            continue
                        nome_chave = normalizar_nome(chave_nome or "")
                        nome_procurado = normalizar_nome(nome_antigo or "")
                        if nome_chave == nome_procurado:
                            reg["apagado"] = True
                            break

    estado_chaves_depois = sorted(list(estado.keys()))
    print(
        "[CLIENTES][SYNC][DEBUG] depois",
        {
            "total_clientes": len(clientes_resultantes),
            "keys": estado_chaves_depois,
        },
    )

    guardar_dados()

    return {
        "ok": True,
        "total_antes": len(clientes_antes),
        "total_depois": len(clientes_resultantes),
        "novos_nifs": sorted(list(nifs_finais - nifs_antes)),
        "removidos_nifs": sorted(list(removidos_nifs)),
    }


# ================== REMOVER CLIENTE ==================

@router.get("/clientes/remover")
async def remover_cliente(idx: int):
    clientes = estado.get("clientes", [])
    if 0 <= idx < len(clientes):
        clientes.pop(idx)
        guardar_dados()
    return RedirectResponse(url="/clientes", status_code=303)


# ================== AUTO-PREENCHIMENTO NIF.pt ==================

@router.post("/clientes/autofill-nif")
async def autofill_nif(nif: str = Form(...)):
    nif = (nif or "").strip()
    if not nif:
        return JSONResponse({"ok": False, "message": "NIF em branco."})

    try:
        url = f"http://www.nif.pt/?json=1&q={nif}&key={API_KEY_NIF}"
        resp = requests.get(url, timeout=5)

        if resp.status_code != 200:
            return JSONResponse(
                {"ok": False, "message": f"Erro HTTP {resp.status_code} ao contactar o NIF.pt."}
            )

        data = resp.json()
        if data.get("result") != "success":
            return JSONResponse(
                {"ok": False, "message": "NIF.pt devolveu um resultado sem sucesso."}
            )

        records = data.get("records", {})
        if not isinstance(records, dict) or not records:
            return JSONResponse(
                {"ok": False, "message": "NIF não encontrado no NIF.pt."}
            )

        rec = next(iter(records.values()))
        nome = rec.get("title") or ""
        morada = rec.get("address") or ""

        if not morada:
            place = rec.get("place", {})
            if isinstance(place, dict):
                morada = place.get("address", "") or ""

        return JSONResponse(
            {
                "ok": True,
                "nif": nif,
                "nome": nome,
                "morada": morada,
            }
        )

    except Exception:
        return JSONResponse(
            {
                "ok": False,
                "message": "Erro ao tentar obter dados no NIF.pt.",
            }
        )
