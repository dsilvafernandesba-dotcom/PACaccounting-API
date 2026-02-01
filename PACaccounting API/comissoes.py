# [VERSÃO ESTÁVEL] Módulo Comissões (cálculo mensal + histórico + exports) — data 2026-01-10

"""PACAccounting comissoes module utilities.

Dependencias para exports: pip install reportlab pillow
"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Tuple
import unicodedata
import json

from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

from dados import estado

router = APIRouter()
templates = Jinja2Templates(directory="templates")

BASE_DIR = Path(__file__).resolve().parent
DATA_FILE = BASE_DIR / "comissoes_dados.json"

ALBERTINA_CARTEIRA = "M Albertina Alves"
ALBERTINA_TARGET_NIFS = {"233884025", "207258120", "208392793", "184169968"}

CANONICAL_CARTEIRAS = [
    "Ana Rodrigues",
    "Celine Santos",
    "Armando Dias",
    ALBERTINA_CARTEIRA,
    "Pedro Fernandes",
]

ALLOWED_CARTEIRAS = set(CANONICAL_CARTEIRAS)

CARTEIRA_ALIASES = {
    "Ana Rodrigues": {"Ana Rodrigues"},
    "Celine Santos": {"Celine Santos"},
    "Armando Dias": {"Armando Dias"},
    ALBERTINA_CARTEIRA: {
        ALBERTINA_CARTEIRA,
        "Maria Albertina Alves",
        "M. Albertina Alves",
        "Albertina Alves",
        "Maria A Alves",
        "Maria A. Alves",
        "Maria Albertina",
        "Maria Albertina Pereira Alves",
        "Maria Albertina P Alves",
        "Maria Albertina P. Alves",
        "Maria Albertina Pereira",
        "Mª Albertina Alves",
        "Mª. Albertina Alves",
    },
    "Pedro Fernandes": {"Pedro Fernandes"},
}

CARTEIRA_FIELD_CANDIDATES = [
    "carteira",
    "Carteira",
    "carteira_nome",
    "carteira_responsavel",
    "carteiraResponsavel",
    "dono_carteira",
    "responsavel_carteira",
    "responsavelCarteira",
    "carteira_responsável",
    "carteiraAtual",
    "carteira_atual",
]

TECNICO_FIELD_CANDIDATES = [
    "tecnico",
    "Técnico",
    "tecnico_principal",
    "tecnicoPrincipal",
    "tecnico_grh",
    "tecnico_responsavel",
    "tecnicoResponsavel",
]

MAX_MENSALIDADES = 12

SCHEMA_VERSION = 2
VERSION_TAG = "stable-2026-01-10"

GOLD_HEX = "#FBBF24"
DARK_HEX = "#020b1f"
GRID_HEX = "#1e3a5f"

_PIL_FONT_PATHS = {
    False: [
        "arial.ttf",
        "Arial.ttf",
        "DejaVuSans.ttf",
    ],
    True: [
        "arialbd.ttf",
        "Arial Bold.ttf",
        "Arial-Bold.ttf",
        "DejaVuSans-Bold.ttf",
    ],
}


def _load_pillow_font(size: int, bold: bool = False):
    paths = _PIL_FONT_PATHS[bool(bold)]
    for path in paths:
        try:
            from PIL import ImageFont  # imported lazily to avoid module cost if unused

            return ImageFont.truetype(path, size)
        except Exception:
            continue
    try:
        from PIL import ImageFont

        return ImageFont.load_default()
    except Exception:
        return None


def _normalize_spaces(text: str) -> str:
    if text is None:
        return ""
    return " ".join(str(text).replace("\xa0", " ").split())


def _carteira_alias_key(valor: str) -> str:
    base = _normalize_spaces(valor)
    if not base:
        return ""
    sem_diacriticos = unicodedata.normalize("NFKD", base)
    sem_diacriticos = "".join(ch for ch in sem_diacriticos if not unicodedata.combining(ch))
    sem_pontuacao = sem_diacriticos.replace(".", " ")
    cleaned = "".join(ch if ch.isalnum() or ch.isspace() else " " for ch in sem_pontuacao.lower())
    return " ".join(cleaned.split())


CARTEIRA_ALIAS_MAP: Dict[str, str] = {}
for canonical, aliases in CARTEIRA_ALIASES.items():
    for alias in set(aliases) | {canonical}:
        key = _carteira_alias_key(alias)
        if key:
            CARTEIRA_ALIAS_MAP[key] = canonical


def _canonical_carteira(valor: str) -> str:
    base = _normalize_spaces(valor)
    if not base:
        return ""
    key = _carteira_alias_key(base)
    if key in CARTEIRA_ALIAS_MAP:
        return CARTEIRA_ALIAS_MAP[key]
    titulo = " ".join(parte[:1].upper() + parte[1:].lower() for parte in base.replace(".", " ").split())
    key_titulo = _carteira_alias_key(titulo)
    return CARTEIRA_ALIAS_MAP.get(key_titulo, titulo)


def _norm_nome(nome: str) -> str:
    nome = _normalize_spaces(nome)
    if not nome:
        return ""
    return " ".join(parte[:1].upper() + parte[1:].lower() for parte in nome.split())


def _cliente_sort_key(item: dict) -> Tuple[str, str, str]:
    carteira_key = _normalize_spaces(item.get("carteira", "")).casefold()
    nome_key = _normalize_spaces(item.get("nome", "")).casefold()
    nif_raw = str(item.get("nif", "")).strip()
    nif_digits = "".join(ch for ch in nif_raw if ch.isdigit()) or nif_raw
    return carteira_key, nome_key, nif_digits


def _extract_carteira_raw(cliente: dict):
    for chave in CARTEIRA_FIELD_CANDIDATES:
        if chave in cliente:
            return cliente.get(chave)
    for chave, valor in cliente.items():
        if isinstance(chave, str) and "carteira" in chave.lower():
            if "tecn" in chave.lower():
                continue
            return valor
    return None


def _parse_euro(valor) -> Decimal:
    if valor is None:
        return Decimal("0")
    if isinstance(valor, Decimal):
        return valor
    if isinstance(valor, (int, float)):
        return Decimal(str(valor))

    texto = str(valor).strip()
    if not texto:
        return Decimal("0")

    texto = texto.replace("€", "").replace(" ", "")
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")

    try:
        return Decimal(texto)
    except InvalidOperation:
        return Decimal("0")


def _fmt_euro(valor: Decimal) -> str:
    valor = (valor or Decimal("0")).quantize(Decimal("0.01"))
    inteiro, decimal = f"{valor:.2f}".split(".")
    blocos = []
    while inteiro:
        blocos.append(inteiro[-3:])
        inteiro = inteiro[:-3]
    inteiro_fmt = ".".join(reversed(blocos)) if blocos else "0"
    return f"{inteiro_fmt},{decimal} €"


def _load_store() -> dict:
    if not DATA_FILE.exists():
        return {}
    try:
        return json.loads(DATA_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _save_store(store: dict) -> None:
    DATA_FILE.write_text(json.dumps(store, ensure_ascii=False, indent=2), encoding="utf-8")


def _get_field(dados: dict, *chaves, default=None):
    for chave in chaves:
        if chave in dados:
            return dados.get(chave)
    return default


def _coerce_bool(valor) -> bool:
    if isinstance(valor, bool):
        return valor
    if isinstance(valor, (int, float)):
        return bool(valor)
    if isinstance(valor, str):
        return valor.strip().lower() in {"1", "true", "sim", "yes", "on"}
    return False


def _safe_int(valor, default: int = 0) -> int:
    try:
        return int(valor)
    except (TypeError, ValueError):
        return default


def _clamp_mensalidades(valor: int) -> int:
    if valor < 0:
        return 0
    if valor > MAX_MENSALIDADES:
        return MAX_MENSALIDADES
    return valor


def _get_clientes_filtrados() -> List[dict]:
    clientes = estado.get("clientes", []) or []
    linhas: List[dict] = []

    for cliente in clientes:
        nif = str(_get_field(cliente, "nif", "NIF", "vat", default="")).strip()
        nome = str(_get_field(cliente, "nome", "Nome", "cliente", default="")).strip()

        carteira_raw = _extract_carteira_raw(cliente)
        carteira = _canonical_carteira(carteira_raw)

        if carteira not in ALLOWED_CARTEIRAS:
            continue

        tecnico_raw = _get_field(cliente, *TECNICO_FIELD_CANDIDATES, default="")
        tecnico_canonical = _canonical_carteira(tecnico_raw)
        tecnico = tecnico_canonical if tecnico_canonical else _norm_nome(tecnico_raw)

        mensalidade = _parse_euro(_get_field(cliente, "mensalidade", "Mensalidade", default=0))
        grh = _parse_euro(_get_field(cliente, "grh", "GRH", default=0))

        taxa = Decimal("0.30") if carteira and tecnico_canonical == carteira else Decimal("0.20")

        linhas.append(
            {
                "nif": nif,
                "nome": nome,
                "carteira": carteira,
                "tecnico": tecnico,
                "mensalidade": mensalidade,
                "grh": grh,
                "taxa": taxa,
            }
        )

    linhas.sort(key=_cliente_sort_key)
    return linhas


def _compose_row(cliente: dict, guardado: dict | None, force_reset: bool = False) -> Tuple[dict, dict, bool]:
    mensalidade = cliente["mensalidade"]
    taxa = cliente["taxa"]

    recebido_flag = False
    num_mensalidades = 0
    needs_update = force_reset

    if not force_reset and guardado:
        if "recebido_mensalidade" in guardado:
            needs_update = True
        elif "recebido" in guardado and "num_mensalidades" in guardado:
            recebido_flag = _coerce_bool(guardado.get("recebido"))
            num_mensalidades = _clamp_mensalidades(_safe_int(guardado.get("num_mensalidades"), 0))
        else:
            needs_update = True

    if not recebido_flag:
        num_mensalidades = 0

    valor_recebido = (mensalidade * num_mensalidades).quantize(Decimal("0.01")) if recebido_flag else Decimal("0")
    comissao = (valor_recebido * taxa).quantize(Decimal("0.01"))

    store_row = {
        "nif": cliente["nif"],
        "nome": cliente["nome"],
        "carteira": cliente["carteira"],
        "tecnico": cliente["tecnico"],
        "taxa": str(taxa),
        "recebido": recebido_flag,
        "num_mensalidades": num_mensalidades,
        "valor_recebido": str(valor_recebido),
        "comissao": str(comissao),
    }

    if guardado and "recebido" in guardado:
        if (
            _coerce_bool(guardado.get("recebido")) != recebido_flag
            or _clamp_mensalidades(_safe_int(guardado.get("num_mensalidades"), 0)) != num_mensalidades
            or _parse_euro(guardado.get("valor_recebido")) != valor_recebido
            or _parse_euro(guardado.get("comissao")) != comissao
        ):
            needs_update = True

    if guardado:
        if guardado.get("carteira") != store_row["carteira"]:
            needs_update = True
        if guardado.get("tecnico") != store_row["tecnico"]:
            needs_update = True
        stored_taxa = guardado.get("taxa")
        if stored_taxa is not None and str(stored_taxa) != store_row["taxa"]:
            needs_update = True

    view_row = {
        **cliente,
        "recebido": recebido_flag,
        "num_mensalidades": num_mensalidades,
        "valor_recebido": valor_recebido,
        "comissao": comissao,
        "taxa_pct": int(taxa * 100),
    }

    return view_row, store_row, needs_update


def _calc_totais(linhas: List[dict]) -> Tuple[Dict[str, Dict[str, Decimal | int]], Dict[str, Decimal | int]]:
    totais_por_carteira: Dict[str, Dict[str, Decimal | int]] = {
        carteira: {"mensalidades": 0, "recebido": Decimal("0"), "comissao": Decimal("0")}
        for carteira in sorted(ALLOWED_CARTEIRAS)
    }
    total_geral: Dict[str, Decimal | int] = {
        "mensalidades": 0,
        "recebido": Decimal("0"),
        "comissao": Decimal("0"),
    }

    for linha in linhas:
        carteira = linha["carteira"]
        if carteira not in totais_por_carteira:
            totais_por_carteira[carteira] = {"mensalidades": 0, "recebido": Decimal("0"), "comissao": Decimal("0")}
        mensalidades = int(linha.get("num_mensalidades", 0) or 0)
        totais_por_carteira[carteira]["mensalidades"] += mensalidades
        totais_por_carteira[carteira]["recebido"] += linha["valor_recebido"]
        totais_por_carteira[carteira]["comissao"] += linha["comissao"]
        total_geral["mensalidades"] += mensalidades
        total_geral["recebido"] += linha["valor_recebido"]
        total_geral["comissao"] += linha["comissao"]

    return totais_por_carteira, total_geral


def _serialize_totals(
    totais_por_carteira: Dict[str, Dict[str, Decimal | int]],
    total_geral: Dict[str, Decimal | int],
) -> dict:
    return {
        "por_carteira": {
            carteira: {
                "mensalidades": int(valores["mensalidades"]),
                "recebido": str(valores["recebido"].quantize(Decimal("0.01"))),
                "comissao": str(valores["comissao"].quantize(Decimal("0.01"))),
            }
            for carteira, valores in totais_por_carteira.items()
        },
        "total_geral": {
            "mensalidades": int(total_geral["mensalidades"]),
            "recebido": str(total_geral["recebido"].quantize(Decimal("0.01"))),
            "comissao": str(total_geral["comissao"].quantize(Decimal("0.01"))),
        },
    }


def _looks_like_auto_prefill(guardados: Dict[str, dict], clientes: List[dict]) -> bool:
    if not guardados:
        return False

    if any("recebido_mensalidade" in dados for dados in guardados.values()):
        return True

    total = len(guardados)
    if total == 0:
        return False

    cliente_por_nif = {cliente["nif"]: cliente for cliente in clientes}
    recebidos = 0
    mensalidade_unica = 0
    valor_coincidente = 0

    for dados in guardados.values():
        if _coerce_bool(dados.get("recebido")):
            recebidos += 1
            if _clamp_mensalidades(_safe_int(dados.get("num_mensalidades"), 0)) == 1:
                mensalidade_unica += 1
            cliente = cliente_por_nif.get(dados.get("nif"))
            if cliente:
                valor_recebido = _parse_euro(dados.get("valor_recebido"))
                mensalidade_cliente = cliente["mensalidade"]
                if mensalidade_cliente > Decimal("0"):
                    if abs(valor_recebido - mensalidade_cliente) <= Decimal("0.05"):
                        valor_coincidente += 1

    if recebidos == 0:
        return False

    proporcao_recebido = recebidos / total
    proporcao_mensalidade = mensalidade_unica / recebidos if recebidos else 0
    proporcao_valores = valor_coincidente / recebidos if recebidos else 0

    return (
        proporcao_recebido >= 0.7
        and proporcao_mensalidade >= 0.7
        and proporcao_valores >= 0.7
    )


def _get_month_rows(
    mes: str,
) -> Tuple[List[dict], Dict[str, Dict[str, Decimal | int]], Dict[str, Decimal | int], str | None]:
    store = _load_store()
    registo = store.get(mes, {}) or {}
    guardados_original = registo.get("rows") or {}
    schema_version = registo.get("schema_version", 1)

    clientes = _get_clientes_filtrados()
    force_reset = schema_version < SCHEMA_VERSION and _looks_like_auto_prefill(guardados_original, clientes)

    guardados = {} if force_reset else dict(guardados_original)
    linhas: List[dict] = []
    needs_save = force_reset

    for cliente in clientes:
        nif = cliente["nif"]
        guardado = guardados_original.get(nif)
        linha, store_row, actualizar = _compose_row(cliente, guardado, force_reset=force_reset)
        linhas.append(linha)
        guardados[nif] = store_row
        if actualizar:
            needs_save = True

    totais_por_carteira, total_geral = _calc_totais(linhas)

    schema_upgrade_needed = (mes in store) and schema_version < SCHEMA_VERSION

    if needs_save or schema_upgrade_needed:
        registo["rows"] = guardados
        registo["totais"] = _serialize_totals(totais_por_carteira, total_geral)
        registo["schema_version"] = SCHEMA_VERSION
        store[mes] = registo
        _save_store(store)

    updated_at = registo.get("updated_at") if registo else None
    return linhas, totais_por_carteira, total_geral, updated_at


def _get_resumo_por_carteira(mes: str) -> Tuple[List[dict], Dict[str, Decimal | int], str | None]:
    _, totais_por_carteira, total_geral, updated_at = _get_month_rows(mes)
    resumo = []
    for carteira in sorted(ALLOWED_CARTEIRAS):
        valores = totais_por_carteira.get(
            carteira,
            {"mensalidades": 0, "recebido": Decimal("0"), "comissao": Decimal("0")},
        )
        resumo.append(
            {
                "carteira": carteira,
                "mensalidades": int(valores["mensalidades"]),
                "recebido": valores["recebido"],
                "comissao": valores["comissao"],
            }
        )
    return resumo, total_geral, updated_at


def _slugify_filename(texto: str) -> str:
    base = "".join(ch if ch.isalnum() else "_" for ch in texto)
    base = base.strip("_") or "comissao"
    return base


def _filtrar_por_carteira(linhas: List[dict], carteira_raw: str) -> Tuple[str, List[dict]]:
    carteira = _canonical_carteira(carteira_raw)
    if not carteira:
        raise HTTPException(status_code=400, detail="Carteira não reconhecida")
    if carteira not in ALLOWED_CARTEIRAS:
        raise HTTPException(status_code=400, detail="Carteira não permitida")
    filtradas = [linha for linha in linhas if linha["carteira"] == carteira]
    return carteira, filtradas


@router.get("/comissoes", response_class=HTMLResponse)
def comissoes_view(request: Request, mes: str | None = None):
    if not mes:
        mes = date.today().strftime("%Y-%m")

    clientes_estado = estado.get("clientes", []) or []
    albertina_count = 0
    nifs_encontrados = set()

    for cliente in clientes_estado:
        carteira_raw = _extract_carteira_raw(cliente)
        carteira_canonica = _canonical_carteira(carteira_raw)
        if carteira_canonica == ALBERTINA_CARTEIRA:
            albertina_count += 1

        nif_valor = str(_get_field(cliente, "nif", "NIF", "vat", default="")).strip()
        nif_digits = "".join(ch for ch in nif_valor if ch.isdigit()) or nif_valor
        if nif_digits in ALBERTINA_TARGET_NIFS:
            nifs_encontrados.add(nif_digits)

    if albertina_count == 0 or not (ALBERTINA_TARGET_NIFS & nifs_encontrados):
        print(
            "[COMISSOES] ATENÇÃO: não encontrei clientes da M Albertina no estado carregado. Verifica se estás a usar o dados.json certo / reinicia o servidor / ficheiro duplicado."
        )

    linhas, totais_por_carteira, total_geral, updated_at = _get_month_rows(mes)

    return templates.TemplateResponse(
        "comissoes.html",
        {
            "request": request,
            "mes": mes,
            "rows": linhas,
            "allowed_carteiras": sorted(ALLOWED_CARTEIRAS),
            "totais_por_carteira": totais_por_carteira,
            "total_geral": total_geral,
            "updated_at": updated_at,
            "fmt_euro": _fmt_euro,
            "max_mensalidades": MAX_MENSALIDADES,
        },
    )


@router.post("/comissoes/guardar")
async def comissoes_guardar(request: Request):
    form = await request.form()
    mes = (form.get("mes") or "").strip() or date.today().strftime("%Y-%m")

    clientes = _get_clientes_filtrados()
    linhas_store: Dict[str, dict] = {}
    linhas_view: List[dict] = []

    for cliente in clientes:
        nif = cliente["nif"]
        recebeu = form.get(f"recebido_{nif}") is not None
        num_raw = form.get(f"num_mensalidades_{nif}")
        num_mensalidades = _clamp_mensalidades(_safe_int(num_raw, 0)) if recebeu else 0

        valor_recebido = (cliente["mensalidade"] * num_mensalidades).quantize(Decimal("0.01")) if recebeu else Decimal("0")
        comissao = (valor_recebido * cliente["taxa"]).quantize(Decimal("0.01"))

        linhas_store[nif] = {
            "nif": nif,
            "nome": cliente["nome"],
            "carteira": cliente["carteira"],
            "tecnico": cliente["tecnico"],
            "taxa": str(cliente["taxa"]),
            "recebido": recebeu,
            "num_mensalidades": num_mensalidades,
            "valor_recebido": str(valor_recebido),
            "comissao": str(comissao),
        }

        linhas_view.append(
            {
                **cliente,
                "recebido": recebeu,
                "num_mensalidades": num_mensalidades,
                "valor_recebido": valor_recebido,
                "comissao": comissao,
                "taxa_pct": int(cliente["taxa"] * 100),
            }
        )

    totais_por_carteira, total_geral = _calc_totais(linhas_view)

    store = _load_store()
    store[mes] = {
        "schema_version": SCHEMA_VERSION,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "module_version": VERSION_TAG,
        "rows": linhas_store,
        "totais": _serialize_totals(totais_por_carteira, total_geral),
    }
    _save_store(store)

    return RedirectResponse(url=f"/comissoes?mes={mes}", status_code=303)


@router.get("/comissoes/exportar-excel")
def comissoes_exportar_excel(mes: str | None = None):
    if not mes:
        mes = date.today().strftime("%Y-%m")

    linhas, _, _, updated_at = _get_month_rows(mes)

    export_rows = []
    for linha in linhas:
        export_rows.append(
            {
                "Carteira": linha["carteira"],
                "Técnico": linha["tecnico"],
                "Cliente": linha["nome"],
                "NIF": linha["nif"],
                "Mensalidade (ref.)": float(linha["mensalidade"]),
                "Recebido?": "Sim" if linha["recebido"] else "Não",
                "Nº mensalidades pagas": linha["num_mensalidades"],
                "Valor recebido (mês)": float(linha["valor_recebido"]),
                "Taxa %": float(linha["taxa"] * 100),
                "Comissão (€)": float(linha["comissao"]),
            }
        )

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        return HTMLResponse(
            "openpyxl não está instalado. Instala com: pip install openpyxl",
            status_code=500,
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Comissões"

    headers = list(export_rows[0].keys()) if export_rows else [
        "Carteira",
        "Técnico",
        "Cliente",
        "NIF",
        "Mensalidade (ref.)",
        "Recebido?",
        "Nº mensalidades pagas",
        "Valor recebido (mês)",
        "Taxa %",
        "Comissão (€)",
    ]
    ws.append(headers)

    for row in export_rows:
        ws.append([row[col] for col in headers])

    for idx, header in enumerate(headers, start=1):
        max_len = max(
            [len(str(header))]
            + [len(str(ws.cell(row=i, column=idx).value or "")) for i in range(2, ws.max_row + 1)]
        )
        ws.column_dimensions[get_column_letter(idx)].width = min(max(12, max_len + 2), 40)

    ws_totais = wb.create_sheet("Sumário")
    ws_totais.append(["Mês", mes])
    ws_totais.append(["Atualizado em", updated_at or "—"])
    ws_totais.append([])
    ws_totais.append(["Carteira", "Mensalidades recebidas", "Valor recebido (€)", "Comissão (€)"])

    totais_por_carteira, total_geral = _calc_totais(linhas)
    for carteira in sorted(ALLOWED_CARTEIRAS):
        totais = totais_por_carteira.get(
            carteira,
            {"mensalidades": 0, "recebido": Decimal("0"), "comissao": Decimal("0")},
        )
        ws_totais.append([
            carteira,
            int(totais.get("mensalidades", 0)),
            float(totais["recebido"]),
            float(totais["comissao"]),
        ])

    ws_totais.append([])
    ws_totais.append([
        "Total geral",
        int(total_geral.get("mensalidades", 0)),
        float(total_geral["recebido"]),
        float(total_geral["comissao"]),
    ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"comissoes_{mes}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/comissoes/exportar-excel-carteira")
def comissoes_exportar_excel_carteira(mes: str | None = None, carteira: str | None = None):
    if not mes:
        mes = date.today().strftime("%Y-%m")
    if not carteira:
        raise HTTPException(status_code=400, detail="Parametro 'carteira' é obrigatório")

    linhas, _, _, _ = _get_month_rows(mes)
    carteira_norm, linhas_carteira = _filtrar_por_carteira(linhas, carteira)

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        return HTMLResponse(
            "openpyxl não está instalado. Instala com: pip install openpyxl",
            status_code=500,
        )

    wb = Workbook()
    ws = wb.active
    ws.title = carteira_norm

    ws.append(["Cliente", "Valor recebido (€)", "Comissão (€)"])

    total_recebido = Decimal("0")
    total_comissao = Decimal("0")

    for linha in linhas_carteira:
        total_recebido += linha["valor_recebido"]
        total_comissao += linha["comissao"]
        ws.append([
            linha["nome"],
            float(linha["valor_recebido"]),
            float(linha["comissao"]),
        ])

    ws.append([])
    ws.append([
        "Total",
        float(total_recebido),
        float(total_comissao),
    ])

    for idx in range(1, 4):
        max_len = max(
            [len(str(ws.cell(row=i, column=idx).value or "")) for i in range(1, ws.max_row + 1)]
        )
        ws.column_dimensions[get_column_letter(idx)].width = min(max(18 if idx == 1 else 14, max_len + 2), 60)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"comissoes_{mes}_{_slugify_filename(carteira_norm)}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/comissoes/exportar-pdf")
def comissoes_exportar_pdf(mes: str | None = None, carteira: str | None = None):
    if not mes:
        mes = date.today().strftime("%Y-%m")
    if not carteira:
        raise HTTPException(status_code=400, detail="Parametro 'carteira' é obrigatório")

    linhas, _, _, _ = _get_month_rows(mes)
    carteira_norm, linhas_carteira = _filtrar_por_carteira(linhas, carteira)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib.colors import HexColor
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    except Exception:
        return HTMLResponse(
            "reportlab não está instalado. Instala com: pip install reportlab",
            status_code=500,
        )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name="CarteiraTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=16,
        textColor=HexColor(GOLD_HEX),
        leading=20,
        spaceAfter=14,
    )

    title = Paragraph(f"Comissões {mes} — {carteira_norm}", title_style)

    data = [["Cliente", "Valor recebido", "Comissão"]]
    total_recebido = Decimal("0")
    total_comissao = Decimal("0")

    for linha in linhas_carteira:
        data.append([
            linha["nome"],
            _fmt_euro(linha["valor_recebido"]),
            _fmt_euro(linha["comissao"]),
        ])
        total_recebido += linha["valor_recebido"]
        total_comissao += linha["comissao"]

    data.append([
        "Total",
        _fmt_euro(total_recebido),
        _fmt_euro(total_comissao),
    ])

    col_widths = [doc.width * 0.5, doc.width * 0.25, doc.width * 0.25]

    tabela = Table(data, colWidths=col_widths, hAlign="LEFT")
    tabela_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HexColor(GOLD_HEX)),
        ("TEXTCOLOR", (0, 0), (-1, 0), HexColor(DARK_HEX)),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("ALIGN", (1, 0), (-1, 0), "CENTER"),
        ("BACKGROUND", (0, -1), (-1, -1), HexColor(GOLD_HEX)),
        ("TEXTCOLOR", (0, -1), (-1, -1), HexColor(DARK_HEX)),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ALIGN", (1, -1), (-1, -1), "CENTER"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("ALIGN", (1, 1), (-1, -2), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, HexColor(GRID_HEX)),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ])

    tabela.setStyle(tabela_style)

    story = [title, Spacer(1, 12), tabela]
    doc.build(story)
    buffer.seek(0)

    filename = f"comissoes_{mes}_{_slugify_filename(carteira_norm)}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/comissoes/exportar-png")
def comissoes_exportar_png(mes: str | None = None, carteira: str | None = None):
    if not mes:
        mes = date.today().strftime("%Y-%m")
    if not carteira:
        raise HTTPException(status_code=400, detail="Parametro 'carteira' é obrigatório")

    linhas, _, _, _ = _get_month_rows(mes)
    carteira_norm, linhas_carteira = _filtrar_por_carteira(linhas, carteira)

    try:
        from PIL import Image, ImageDraw, ImageFont
    except Exception:
        return HTMLResponse(
            "Pillow não está instalado. Instala com: pip install pillow",
            status_code=500,
        )

    linhas_count = max(len(linhas_carteira), 1)
    linha_altura = 46
    largura = 900
    margem_esquerda = 40
    margem_superior = 60
    altura = margem_superior + (linhas_count + 5) * linha_altura

    imagem = Image.new("RGB", (largura, altura), "#020b1f")
    draw = ImageDraw.Draw(imagem)

    def carregar_fonte(tamanho: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
        try:
            return ImageFont.truetype("arial.ttf", tamanho)
        except Exception:
            try:
                return ImageFont.truetype("DejaVuSans.ttf", tamanho)
            except Exception:
                return ImageFont.load_default()

    fonte_titulo = carregar_fonte(28)
    fonte_cabecalho = carregar_fonte(18)
    fonte_texto = carregar_fonte(16)

    draw.text((margem_esquerda, 20), f"Comissões {mes} - {carteira_norm}", fill="#fbbf24", font=fonte_titulo)

    y = margem_superior
    draw.text((margem_esquerda, y), "Cliente", fill="#f9fafb", font=fonte_cabecalho)
    draw.text((margem_esquerda + 380, y), "Valor recebido", fill="#f9fafb", font=fonte_cabecalho)
    draw.text((margem_esquerda + 620, y), "Comissão", fill="#f9fafb", font=fonte_cabecalho)
    y += linha_altura

    total_recebido = Decimal("0")
    total_comissao = Decimal("0")

    for linha in linhas_carteira:
        draw.text((margem_esquerda, y), linha["nome"][:40], fill="#f9fafb", font=fonte_texto)
        draw.text((margem_esquerda + 380, y), _fmt_euro(linha["valor_recebido"]), fill="#fbbf24", font=fonte_texto)
        draw.text((margem_esquerda + 620, y), _fmt_euro(linha["comissao"]), fill="#fbbf24", font=fonte_texto)
        total_recebido += linha["valor_recebido"]
        total_comissao += linha["comissao"]
        y += linha_altura

    y += 10
    draw.line((margem_esquerda, y, largura - margem_esquerda, y), fill="#fbbf24", width=2)
    y += 20

    draw.text((margem_esquerda, y), "Total", fill="#fbbf24", font=fonte_cabecalho)
    draw.text((margem_esquerda + 380, y), _fmt_euro(total_recebido), fill="#fbbf24", font=fonte_cabecalho)
    draw.text((margem_esquerda + 620, y), _fmt_euro(total_comissao), fill="#fbbf24", font=fonte_cabecalho)

    buffer = BytesIO()
    imagem.save(buffer, format="PNG")
    buffer.seek(0)

    filename = f"comissoes_{mes}_{_slugify_filename(carteira_norm)}.png"
    return StreamingResponse(
        buffer,
        media_type="image/png",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/comissoes/exportar-resumo-excel")
def comissoes_exportar_resumo_excel(mes: str | None = None):
    if not mes:
        mes = date.today().strftime("%Y-%m")

    resumo, total_geral, updated_at = _get_resumo_por_carteira(mes)

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:
        return HTMLResponse(
            "openpyxl não está instalado. Instala com: pip install openpyxl",
            status_code=500,
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"

    ws.append(["Carteira", "Mensalidades", "Valor recebido", "Comissão"])
    for linha in resumo:
        ws.append([
            linha["carteira"],
            int(linha.get("mensalidades", 0)),
            float(linha["recebido"]),
            float(linha["comissao"]),
        ])

    ws.append([])
    ws.append([
        "Total geral",
        int(total_geral.get("mensalidades", 0)),
        float(total_geral["recebido"]),
        float(total_geral["comissao"]),
    ])

    for idx in range(1, 5):
        max_len = max(
            [len(str(ws.cell(row=i, column=idx).value or "")) for i in range(1, ws.max_row + 1)]
        )
        ws.column_dimensions[get_column_letter(idx)].width = min(max(18 if idx == 1 else 14, max_len + 2), 60)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10233F", end_color="10233F", fill_type="solid")
    bold_font = Font(bold=True)
    footer_fill = PatternFill(start_color="10233F", end_color="10233F", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")

    for idx, cell in enumerate(ws[1], start=1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = left_alignment if idx == 1 else center_alignment

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[0].value is None:
            continue
        is_total = str(row[0].value).strip().lower() == "total geral"
        if is_total:
            for cell in row:
                cell.font = header_font
                cell.fill = footer_fill
                cell.alignment = center_alignment
            row[0].alignment = left_alignment
        else:
            row[0].font = bold_font
            row[0].alignment = left_alignment
            for cell in row[1:]:
                cell.font = Font(bold=False)
                cell.alignment = center_alignment
        if row[1].value is not None:
            row[1].number_format = "#,##0"
        if row[2].value is not None:
            row[2].number_format = '#,##0.00 "€"'
        if row[3].value is not None:
            row[3].number_format = '#,##0.00 "€"'

    ws_meta = wb.create_sheet("Metadados")
    ws_meta.append(["Mês", mes])
    ws_meta.append(["Atualizado em", updated_at or "—"])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"resumo_comissoes_{mes}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/comissoes/exportar-resumo-pdf")
def comissoes_exportar_resumo_pdf(mes: str | None = None):
    if not mes:
        mes = date.today().strftime("%Y-%m")

    resumo, total_geral, _ = _get_resumo_por_carteira(mes)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib.colors import HexColor
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    except Exception:
        return HTMLResponse(
            "reportlab não está instalado. Instala com: pip install reportlab",
            status_code=500,
        )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name="ResumoTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=16,
        textColor=HexColor(GOLD_HEX),
        leading=20,
        spaceAfter=14,
    )

    title = Paragraph(f"Resumo Comissões {mes}", title_style)

    data = [["Carteira", "Mensalidades", "Valor recebido", "Comissão"]]
    for linha in resumo:
        data.append([
            linha["carteira"],
            str(int(linha.get("mensalidades", 0))),
            _fmt_euro(linha["recebido"]),
            _fmt_euro(linha["comissao"]),
        ])

    data.append([
        "Total geral",
        str(int(total_geral.get("mensalidades", 0))),
        _fmt_euro(total_geral["recebido"]),
        _fmt_euro(total_geral["comissao"]),
    ])

    col_widths = [doc.width * 0.34, doc.width * 0.18, doc.width * 0.24, doc.width * 0.24]

    tabela = Table(data, colWidths=col_widths, hAlign="LEFT")
    tabela_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HexColor(GOLD_HEX)),
        ("TEXTCOLOR", (0, 0), (-1, 0), HexColor(DARK_HEX)),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("ALIGN", (1, 0), (-1, 0), "CENTER"),
        ("BACKGROUND", (0, -1), (-1, -1), HexColor(GOLD_HEX)),
        ("TEXTCOLOR", (0, -1), (-1, -1), HexColor(DARK_HEX)),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ALIGN", (1, -1), (-1, -1), "CENTER"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("ALIGN", (1, 1), (-1, -2), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, HexColor(GRID_HEX)),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ])

    tabela.setStyle(tabela_style)

    story = [title, Spacer(1, 12), tabela]
    doc.build(story)
    buffer.seek(0)

    filename = f"resumo_comissoes_{mes}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/comissoes/exportar-resumo-png")
def comissoes_exportar_resumo_png(mes: str | None = None):
    if not mes:
        mes = date.today().strftime("%Y-%m")

    resumo, total_geral, _ = _get_resumo_por_carteira(mes)

    try:
        from PIL import Image, ImageDraw
    except Exception:
        return HTMLResponse(
            "Pillow não está instalado. Instala com: pip install pillow",
            status_code=500,
        )

    linhas_count = max(len(resumo), 1)
    margem_horizontal = 64
    margem_vertical = 92
    header_altura = 70
    linha_altura = 58
    footer_altura = 70

    colunas = [
        {"titulo": "Carteira", "largura": 260, "align": "left"},
        {"titulo": "Mensalidades", "largura": 140, "align": "center"},
        {"titulo": "Valor recebido", "largura": 160, "align": "center"},
        {"titulo": "Comissão", "largura": 160, "align": "center"},
    ]

    tabela_largura = sum(col["largura"] for col in colunas)
    largura = tabela_largura + margem_horizontal * 2
    altura = margem_vertical + header_altura + linhas_count * linha_altura + footer_altura + margem_vertical

    imagem = Image.new("RGB", (largura, altura), DARK_HEX)
    draw = ImageDraw.Draw(imagem)

    fonte_titulo = _load_pillow_font(36, bold=True)
    fonte_header = _load_pillow_font(22, bold=True)
    fonte_texto = _load_pillow_font(18, bold=False)
    fonte_texto_bold = _load_pillow_font(18, bold=True)

    texto_claro = "#f9fafb"

    col_posicoes: List[Tuple[int, int]] = []
    cursor_x = margem_horizontal
    for col in colunas:
        col_posicoes.append((cursor_x, cursor_x + col["largura"]))
        cursor_x += col["largura"]

    def desenhar_texto(texto: str, coluna_idx: int, topo: int, altura_celula: int, fonte, align: str, cor: str) -> None:
        inicio, fim = col_posicoes[coluna_idx]
        largura_celula = fim - inicio
        try:
            bbox = draw.textbbox((0, 0), texto, font=fonte)
            texto_largura = bbox[2] - bbox[0]
            texto_altura = bbox[3] - bbox[1]
        except AttributeError:
            texto_largura, texto_altura = draw.textsize(texto, font=fonte)

        if align == "center":
            pos_x = inicio + (largura_celula - texto_largura) / 2
        elif align == "right":
            pos_x = fim - texto_largura - 12
        else:
            pos_x = inicio + 12

        pos_y = topo + (altura_celula - texto_altura) / 2
        draw.text((pos_x, pos_y), texto, fill=cor, font=fonte)

    titulo_y = margem_vertical - 54
    draw.text((margem_horizontal, titulo_y), f"Resumo Comissões {mes}", fill=GOLD_HEX, font=fonte_titulo)

    header_topo = margem_vertical
    header_base = header_topo + header_altura
    draw.rectangle([margem_horizontal, header_topo, margem_horizontal + tabela_largura, header_base], fill=GOLD_HEX)

    for idx, coluna in enumerate(colunas):
        desenhar_texto(coluna["titulo"], idx, header_topo, header_altura, fonte_header, coluna["align"], DARK_HEX)

    corpo_topo = header_base
    y_atual = corpo_topo

    for linha in resumo:
        desenhar_texto(linha["carteira"], 0, y_atual, linha_altura, fonte_texto_bold, colunas[0]["align"], texto_claro)
        desenhar_texto(str(int(linha.get("mensalidades", 0))), 1, y_atual, linha_altura, fonte_texto, colunas[1]["align"], texto_claro)
        desenhar_texto(_fmt_euro(linha["recebido"]), 2, y_atual, linha_altura, fonte_texto, colunas[2]["align"], texto_claro)
        desenhar_texto(_fmt_euro(linha["comissao"]), 3, y_atual, linha_altura, fonte_texto, colunas[3]["align"], texto_claro)

        y_atual += linha_altura
        draw.line((margem_horizontal, y_atual, margem_horizontal + tabela_largura, y_atual), fill=GRID_HEX, width=1)

    footer_topo = corpo_topo + linhas_count * linha_altura
    footer_base = footer_topo + footer_altura
    draw.rectangle([margem_horizontal, footer_topo, margem_horizontal + tabela_largura, footer_base], fill=GOLD_HEX)

    desenhar_texto("Total geral", 0, footer_topo, footer_altura, fonte_header, colunas[0]["align"], DARK_HEX)
    desenhar_texto(str(int(total_geral.get("mensalidades", 0))), 1, footer_topo, footer_altura, fonte_header, colunas[1]["align"], DARK_HEX)
    desenhar_texto(_fmt_euro(total_geral["recebido"]), 2, footer_topo, footer_altura, fonte_header, colunas[2]["align"], DARK_HEX)
    desenhar_texto(_fmt_euro(total_geral["comissao"]), 3, footer_topo, footer_altura, fonte_header, colunas[3]["align"], DARK_HEX)

    buffer = BytesIO()
    imagem.save(buffer, format="PNG")
    buffer.seek(0)

    filename = f"resumo_comissoes_{mes}.png"
    return StreamingResponse(
        buffer,
        media_type="image/png",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
