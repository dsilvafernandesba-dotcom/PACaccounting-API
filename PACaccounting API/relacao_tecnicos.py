# VERSÃO ESTÁVEL: Relação Técnicos (matching por tokens + export Excel/PDF com layout) – 2026-01-07
from __future__ import annotations

import difflib
import json
import os
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import urlencode

from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as XLTable, TableStyleInfo

from dados import estado

router = APIRouter()
templates = Jinja2Templates(directory="templates")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TIMINGS_FILE = os.path.join(BASE_DIR, "timings_dados.json")

# ===== Configuração =====
VALOR_HORA_EUR_DEFAULT = 40.0


# ===== Estilo export =====
_GOLD = "D4AF37"
_DARK_BLUE = "061A44"
_LIGHT_GRAY = "F3F4F6"
_WHITE = "FFFFFF"
_ALERT_RED = "FEE2E2"
_ALERT_AMBER = "FEF3C7"

QUALITY_FLAGS = {
    "sem_tecnico": "Sem técnico",
    "sem_tipo": "Sem tipo contabilidade",
    "sem_periodicidade": "Sem periodicidade IVA",
    "sem_regime": "Sem regime IVA",
    "sem_timings": "Sem timings",
}

STOP_TOKENS = {
    "LDA",
    "LTD",
    "LTDA",
    "SA",
    "SOCIEDADE",
    "UNIPESSOAL",
    "UNIP",
    "E",
    "ME",
    "LDA.",
    "LDA,",
    "LDA;",
    "UNIPESSOAL,",
    "UNIPESSOAL;",
}


def normalizar_nome(valor: Optional[str]) -> str:
    if valor is None:
        return ""
    texto = unicodedata.normalize("NFD", str(valor).strip())
    texto = "".join(ch for ch in texto if unicodedata.category(ch) != "Mn")
    texto = texto.upper()
    texto = re.sub(r"[^A-Z0-9]+", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def tokens_relevantes(norm: str) -> List[str]:
    toks = norm.split()
    out = []
    for t in toks:
        t = t.strip()
        if not t or t in STOP_TOKENS:
            continue
        out.append(t)
    return out


def _format_minutos(total_minutos: int) -> str:
    total_minutos = max(int(total_minutos or 0), 0)
    horas = total_minutos // 60
    mins = total_minutos % 60
    return f"{horas}h{mins:02d}m"


def _parse_minutos(v: Any) -> int:
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(round(float(v)))
    s = str(v).strip()
    if not s:
        return 0
    try:
        if "h" in s.lower() or "m" in s.lower():
            h = 0.0
            m = 0
            mh = re.search(r"([0-9]+(?:[.,][0-9]+)?)\s*h", s.lower())
            mm = re.search(r"([0-9]+)\s*m", s.lower())
            if mh:
                h = float(mh.group(1).replace(",", "."))
            if mm:
                m = int(mm.group(1))
            return int(round(h * 60)) + m
        if ":" in s:
            p = s.split(":")
            if len(p) == 2:
                return int(p[0]) * 60 + int(p[1])
        return int(round(float(s.replace(",", "."))))
    except Exception:
        return 0


def _load_timings() -> Dict[str, Any]:
    if not os.path.exists(TIMINGS_FILE):
        return {}
    try:
        with open(TIMINGS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


@dataclass
class ClienteRow:
    nome: str
    nif: str
    tecnico: str
    tecnico_grh: str
    mensalidade_eur: float
    tipo_contabilidade: str
    periodicidade_iva: str
    regime_iva: str
    timing_media_minutos: int
    timing_media_str: str
    qualidade: List[str]
    detalhe_href: Optional[str]
    primary_sort_key: str
    secondary_sort_key: str
    search_blob: str
    origem_match: str
    chave_timings: Optional[str]
    sem_timings_tooltip: str


# Cache de índice por id(ano_dict)
_MATCH_CACHE: Dict[int, Tuple[Any, Dict[str, List[str]], Dict[str, Dict[str, Any]], List[str], Dict[str, List[str]]]] = {}
# tuple: (ref, map_norm, original_map, norm_keys, tokens_por_norm)


def _construir_indice_timings(
    ano_dict: Dict[str, Any]
) -> Tuple[Dict[str, List[str]], Dict[str, Dict[str, Any]], List[str], Dict[str, List[str]]]:
    map_norm_to_originals: Dict[str, List[str]] = {}
    original_para_registo: Dict[str, Dict[str, Any]] = {}
    tokens_por_norm: Dict[str, List[str]] = {}

    for empresa, rec in ano_dict.items():
        if not isinstance(rec, dict):
            continue
        if rec.get("apagado"):
            continue

        original_para_registo[empresa] = rec
        chave_norm = normalizar_nome(empresa)
        if not chave_norm:
            continue
        map_norm_to_originals.setdefault(chave_norm, []).append(empresa)
        if chave_norm not in tokens_por_norm:
            tokens_por_norm[chave_norm] = tokens_relevantes(chave_norm)

    norm_keys = list(map_norm_to_originals.keys())
    return map_norm_to_originals, original_para_registo, norm_keys, tokens_por_norm


def _obter_match_cache(
    ano_dict: Dict[str, Any]
) -> Tuple[Dict[str, List[str]], Dict[str, Dict[str, Any]], List[str], Dict[str, List[str]]]:
    key = id(ano_dict)
    cached = _MATCH_CACHE.get(key)
    if cached and cached[0] is ano_dict:
        return cached[1], cached[2], cached[3], cached[4]

    map_norm, original_map, norm_keys, tokens_por_norm = _construir_indice_timings(ano_dict)
    _MATCH_CACHE.clear()
    _MATCH_CACHE[key] = (ano_dict, map_norm, original_map, norm_keys, tokens_por_norm)
    return map_norm, original_map, norm_keys, tokens_por_norm


def _escolher_original(nome_cliente: str, candidatos: List[str]) -> Optional[str]:
    if not candidatos:
        return None
    if len(candidatos) == 1:
        return candidatos[0]
    alvo = (nome_cliente or "").strip().upper()
    melhor = candidatos[0]
    melhor_ratio = -1.0
    for c in candidatos:
        r = difflib.SequenceMatcher(None, alvo, c.strip().upper()).ratio()
        if r > melhor_ratio:
            melhor_ratio = r
            melhor = c
    return melhor


def _format_debug_tooltip(cands: List[Tuple[str, float]]) -> str:
    if not cands:
        return ""
    parts = [f"{n} ({sc:.2f})" for n, sc in cands if n]
    return "Sugestões: " + "; ".join(parts) if parts else ""


def match_timings(
    nome_cliente: str,
    ano_dict: Dict[str, Any],
) -> Tuple[Optional[Dict[str, Any]], str, Optional[str], List[Tuple[str, float]]]:
    """
    exato -> inclusao_tokens -> fuzzy (limitado) -> none
    devolve debug_top3 só quando falha
    """
    if not isinstance(ano_dict, dict) or not ano_dict:
        return None, "none", None, []

    nome_original = (nome_cliente or "").strip()
    nome_norm = normalizar_nome(nome_original)
    if not nome_norm:
        return None, "none", None, []

    map_norm, original_map, norm_keys, tokens_por_norm = _obter_match_cache(ano_dict)

    # 1) Exato
    bucket = map_norm.get(nome_norm)
    if bucket:
        chosen = _escolher_original(nome_original, bucket)
        if chosen and chosen in original_map:
            return original_map[chosen], "exact", chosen, []

    # 2) Inclusão por tokens relevantes
    nome_tokens = set(tokens_relevantes(nome_norm))
    candidatos_inclusao: List[Tuple[str, float, int]] = []
    if nome_tokens:
        for chave_norm in norm_keys:
            toks = tokens_por_norm.get(chave_norm, [])
            if not toks:
                continue
            chave_tokens = set(toks)
            if nome_tokens.issubset(chave_tokens) or chave_tokens.issubset(nome_tokens):
                overlap = len(nome_tokens & chave_tokens)
                # ratio só para desempate
                ratio = difflib.SequenceMatcher(None, nome_norm, chave_norm).ratio()
                candidatos_inclusao.append((chave_norm, ratio, overlap))
    if candidatos_inclusao:
        chave_escolhida = max(candidatos_inclusao, key=lambda x: (x[2], x[1], len(x[0])))[0]
        chosen = _escolher_original(nome_original, map_norm[chave_escolhida])
        if chosen and chosen in original_map:
            return original_map[chosen], "inclusao", chosen, []

    # 3) Fuzzy limitado (para evitar “pendurar” em datasets grandes)
    # calcula ratios só para um subconjunto plausível:
    # - primeiro tenta get_close_matches com cutoff
    debug_top3: List[Tuple[str, float]] = []
    try:
        close = difflib.get_close_matches(nome_norm, norm_keys, n=20, cutoff=0.80)
    except Exception:
        close = []

    # se não houver close, limita a 200 primeiros para debug (não mais)
    candidatos_fuzzy = close if close else norm_keys[:200]

    ratios: List[Tuple[str, float]] = []
    for k in candidatos_fuzzy:
        ratios.append((k, difflib.SequenceMatcher(None, nome_norm, k).ratio()))
    ratios.sort(key=lambda x: x[1], reverse=True)

    for k, sc in ratios[:3]:
        bucket_k = map_norm.get(k, [])
        chosen_k = _escolher_original(nome_original, bucket_k) or (bucket_k[0] if bucket_k else k)
        debug_top3.append((chosen_k, sc))

    if ratios:
        best_k, best_sc = ratios[0]
        second_sc = ratios[1][1] if len(ratios) > 1 else 0.0
        if best_sc >= 0.92 and (len(ratios) == 1 or (best_sc - second_sc) >= 0.03):
            chosen = _escolher_original(nome_original, map_norm.get(best_k, []))
            if chosen and chosen in original_map:
                return original_map[chosen], "fuzzy", chosen, []

    return None, "none", None, debug_top3


def _calcular_media_mensal(rec: Optional[Dict[str, Any]]) -> int:
    """média mensal (min) = (Σ meses/12) + extra_mensal"""
    if not isinstance(rec, dict):
        return 0
    meses_dict = rec.get("meses", {}) if isinstance(rec.get("meses"), dict) else {}
    total = 0
    # aceita chaves 1..12 (int) ou "1".."12"
    for mes in range(1, 13):
        raw = meses_dict.get(mes)
        if raw is None:
            raw = meses_dict.get(str(mes))
        total += max(_parse_minutos(raw), 0)
    extra = max(_parse_minutos(rec.get("extra_mensal")), 0)
    return max(int(round(total / 12.0 + extra)), 0)


def _coletar_clientes() -> List[Dict[str, Any]]:
    clientes_raw = estado.get("clientes", [])
    if not isinstance(clientes_raw, list):
        return []
    out: List[Dict[str, Any]] = []
    for idx, c in enumerate(clientes_raw):
        if isinstance(c, dict):
            d = c.copy()
            d["_idx"] = idx
            out.append(d)
    return out


def _build_rows(clientes: Iterable[Dict[str, Any]], ano_dict: Dict[str, Any]) -> List[ClienteRow]:
    rows: List[ClienteRow] = []
    for c in clientes:
        nome = str(c.get("nome") or "").strip()
        nif = str(c.get("nif") or "").strip()
        tecnico = str(c.get("tecnico") or "").strip()
        tecnico_grh = str(c.get("tecnico_grh") or "").strip()
        mensalidade_raw = c.get("mensalidade") or c.get("mensalidade_ref") or c.get("mensalidade_base")
        mensalidade_eur = 0.0
        if isinstance(mensalidade_raw, (int, float)):
            mensalidade_eur = float(mensalidade_raw)
        elif isinstance(mensalidade_raw, str):
            texto = mensalidade_raw.strip().replace("€", "").replace(" ", "")
            if texto:
                texto = texto.replace(".", "").replace(",", ".") if texto.count(",") == 1 and texto.count(".") > 1 else texto.replace(".", "").replace(",", ".")
                try:
                    mensalidade_eur = float(texto)
                except Exception:
                    mensalidade_eur = 0.0
        tipo_cont = str(c.get("tipo_contabilidade") or "").strip()
        periodicidade_iva = str(c.get("periodicidade_iva") or "").strip()
        regime_iva = str(c.get("regime_iva") or "").strip()

        rec, origem, chave, debug = match_timings(nome, ano_dict)
        media_min = _calcular_media_mensal(rec) if rec else 0
        media_str = _format_minutos(media_min)
        tooltip = _format_debug_tooltip(debug) if rec is None else ""

        qualidade: List[str] = []
        if not tecnico:
            qualidade.append(QUALITY_FLAGS["sem_tecnico"])
        if not tipo_cont:
            qualidade.append(QUALITY_FLAGS["sem_tipo"])
        if not periodicidade_iva:
            qualidade.append(QUALITY_FLAGS["sem_periodicidade"])
        if not regime_iva:
            qualidade.append(QUALITY_FLAGS["sem_regime"])
        if rec is None:
            qualidade.append(QUALITY_FLAGS["sem_timings"])

        detalhe_href = None
        idx = c.get("_idx")
        if isinstance(idx, int):
            detalhe_href = f"/clientes/editar/{idx}"

        search_blob = " ".join([nome, nif, tecnico, tecnico_grh]).casefold()
        primary_key = tecnico.casefold() if tecnico else "zzzz"
        secondary_key = nome.casefold()

        rows.append(
            ClienteRow(
                nome=nome,
                nif=nif,
                tecnico=tecnico,
                tecnico_grh=tecnico_grh,
                mensalidade_eur=mensalidade_eur,
                tipo_contabilidade=tipo_cont,
                periodicidade_iva=periodicidade_iva,
                regime_iva=regime_iva,
                timing_media_minutos=media_min,
                timing_media_str=media_str,
                qualidade=qualidade,
                detalhe_href=detalhe_href,
                primary_sort_key=primary_key,
                secondary_sort_key=secondary_key,
                search_blob=search_blob,
                origem_match=origem,
                chave_timings=chave,
                sem_timings_tooltip=tooltip,
            )
        )
    return rows


def _filtrar_rows(
    rows: Iterable[ClienteRow],
    search: str,
    tecnico: str,
    tipo_cont: str,
    periodicidade: str,
    regime: str,
) -> List[ClienteRow]:
    termo = (search or "").casefold()
    out: List[ClienteRow] = []
    for r in rows:
        if termo and termo not in r.search_blob:
            continue
        if tecnico and r.tecnico != tecnico:
            continue
        if tipo_cont and r.tipo_contabilidade != tipo_cont:
            continue
        if periodicidade and r.periodicidade_iva != periodicidade:
            continue
        if regime and r.regime_iva != regime:
            continue
        out.append(r)
    return out


def _ordenar_rows(rows: List[ClienteRow], sort: str, direction: str) -> List[ClienteRow]:
    reverse = direction == "desc"
    if sort == "nome":
        return sorted(rows, key=lambda r: (r.secondary_sort_key, r.timing_media_minutos), reverse=reverse)
    if sort == "timing":
        return sorted(rows, key=lambda r: (r.timing_media_minutos, r.secondary_sort_key), reverse=reverse)

    com_tecnico = [r for r in rows if r.tecnico]
    sem_tecnico = [r for r in rows if not r.tecnico]
    com_tecnico = sorted(com_tecnico, key=lambda r: (r.primary_sort_key, r.secondary_sort_key, r.timing_media_minutos), reverse=reverse)
    sem_tecnico = sorted(sem_tecnico, key=lambda r: (r.secondary_sort_key, r.timing_media_minutos))
    return com_tecnico + sem_tecnico


def _filter_options(rows: Iterable[ClienteRow]) -> Dict[str, List[str]]:
    tecnicos = sorted({r.tecnico for r in rows if r.tecnico})
    tipos = sorted({r.tipo_contabilidade for r in rows if r.tipo_contabilidade})
    periodicidades = sorted({r.periodicidade_iva for r in rows if r.periodicidade_iva})
    regimes = sorted({r.regime_iva for r in rows if r.regime_iva})
    return {"tecnicos": tecnicos, "tipos": tipos, "periodicidades": periodicidades, "regimes": regimes}


def _dataset(request: Request) -> Dict[str, Any]:
    timings_all = _load_timings()
    anos_disponiveis = sorted(int(k) for k in timings_all.keys() if isinstance(k, str) and k.isdigit())
    ano_atual = date.today().year
    if ano_atual not in anos_disponiveis:
        anos_disponiveis.append(ano_atual)
        anos_disponiveis.sort()

    q = request.query_params
    ano_sel: Optional[int] = None
    ano_bruto = q.get("ano")
    if ano_bruto:
        try:
            ano_sel = int(ano_bruto)
        except Exception:
            ano_sel = None
    if not ano_sel:
        ano_sel = ano_atual if ano_atual else (anos_disponiveis[-1] if anos_disponiveis else None)

    ano_dict = timings_all.get(str(ano_sel), {}) if ano_sel else {}
    if not isinstance(ano_dict, dict):
        ano_dict = {}

    clientes = _coletar_clientes()
    rows_base = _build_rows(clientes, ano_dict)
    opcoes = _filter_options(rows_base)

    search = (q.get("search") or "").strip()
    filtro_tecnico = q.get("tecnico") or ""
    filtro_tipo = q.get("tipo_contabilidade") or ""
    filtro_periodicidade = q.get("periodicidade_iva") or ""
    filtro_regime = q.get("regime_iva") or ""

    rows_filtrados = _filtrar_rows(rows_base, search, filtro_tecnico, filtro_tipo, filtro_periodicidade, filtro_regime)

    sort_by = q.get("sort") or "tecnico"
    direction = q.get("dir") or "asc"
    if direction not in {"asc", "desc"}:
        direction = "asc"
    if sort_by not in {"nome", "tecnico", "timing"}:
        sort_by = "tecnico"

    rows = _ordenar_rows(rows_filtrados, sort_by, direction)

    total_min = sum(r.timing_media_minutos for r in rows)
    total_str = _format_minutos(total_min)

    base_params = {
        "search": search,
        "tecnico": filtro_tecnico,
        "tipo_contabilidade": filtro_tipo,
        "periodicidade_iva": filtro_periodicidade,
        "regime_iva": filtro_regime,
        "sort": sort_by,
        "dir": direction,
    }
    if ano_sel:
        base_params["ano"] = str(ano_sel)

    sort_links: Dict[str, str] = {}
    for campo in ["nome", "tecnico", "timing"]:
        params = base_params.copy()
        if campo == sort_by:
            params["dir"] = "desc" if direction == "asc" else "asc"
        else:
            params["dir"] = "asc"
        params["sort"] = campo
        sort_links[campo] = f"/relacao-tecnicos?{urlencode(params)}"

    export_qs = urlencode(base_params)
    base_params_sem_tecnico = base_params.copy()
    base_params_sem_tecnico.pop("tecnico", None)
    export_qs_sem_tecnico = urlencode(base_params_sem_tecnico)

    blocos_tecnico: List[Dict[str, Any]] = []
    blocos_index: Dict[str, Dict[str, Any]] = {}

    def _obter_bloco(tecnico_nome: str) -> Dict[str, Any]:
        chave = _slugify_tecnico_filename(tecnico_nome or "sem_tecnico")
        bloco = blocos_index.get(chave)
        if not bloco:
            bloco = {
                "nome": tecnico_nome.strip() if tecnico_nome else "",
                "slug": chave,
                "rows": [],
            }
            blocos_index[chave] = bloco
            blocos_tecnico.append(bloco)
        return bloco

    for row in rows:
        nome_limpo = (row.tecnico or "").strip()
        bloco_destino = _obter_bloco(nome_limpo)
        bloco_destino["rows"].append(row)

    for idx, bloco in enumerate(blocos_tecnico, start=1):
        bloco["index"] = idx

    tecnicos_lista = [
        {"nome": bloco["nome"], "slug": bloco["slug"], "index": bloco["index"]}
        for bloco in blocos_tecnico
        if bloco["nome"]
    ]

    return {
        "rows": rows,
        "rows_base": rows_base,
        "filtros": {
            "search": search,
            "tecnico": filtro_tecnico,
            "tipo_contabilidade": filtro_tipo,
            "periodicidade_iva": filtro_periodicidade,
            "regime_iva": filtro_regime,
        },
        "opcoes": opcoes,
        "sort_by": sort_by,
        "direction": direction,
        "sort_links": sort_links,
        "total_str": total_str,
        "export_qs": export_qs,
        "export_qs_sem_tecnico": export_qs_sem_tecnico,
        "ano_sel": ano_sel,
        "anos_disponiveis": anos_disponiveis,
        "contagem": len(rows),
        "tecnico_blocks": blocos_tecnico,
        "tecnicos_lista": tecnicos_lista,
    }


# =========================
# Helpers export técnico
# =========================

def _slugify_tecnico_filename(nome: str) -> str:
    texto = unicodedata.normalize("NFKD", nome or "")
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r"[^0-9A-Za-z]+", "_", texto)
    texto = texto.strip("_").lower()
    return texto or "tecnico"


def _prepare_tecnico_export(
    request: Request,
    tecnico_raw: str,
    valor_hora: float,
) -> Tuple[List[ClienteRow], str, Dict[str, str], Optional[int], str]:
    if tecnico_raw is None:
        raise HTTPException(status_code=400, detail="Parametro 'tecnico' é obrigatório")

    tecnico_clean = (tecnico_raw or "").strip()
    if not tecnico_clean:
        raise HTTPException(status_code=400, detail="Parametro 'tecnico' é obrigatório")

    dados = _dataset(request)
    alvo_cf = tecnico_clean.casefold()
    rows_visiveis: List[ClienteRow] = dados.get("rows", []) or []
    linhas_tecnico = [
        r for r in rows_visiveis if (r.tecnico or "").strip().casefold() == alvo_cf
    ]

    tecnico_display = linhas_tecnico[0].tecnico if linhas_tecnico else tecnico_clean
    total_min = sum(r.timing_media_minutos for r in linhas_tecnico)
    total_str = _format_minutos(total_min)

    filtros_export = dict(dados.get("filtros", {}))
    filtros_export["tecnico"] = tecnico_display
    filtros_export["valor_hora"] = f"{valor_hora:.2f}"

    return linhas_tecnico, total_str, filtros_export, dados.get("ano_sel"), tecnico_display


# =========================
# Export Excel bonito
# =========================

def _format_horas_minutos(minutos: int) -> str:
    minutos = int(round(minutos)) if minutos is not None else 0
    minutos = max(minutos, 0)
    h = minutos // 60
    m = minutos % 60
    return f"{h}h{m:02d}m"


def _calcular_tempo_limites(row: ClienteRow, valor_hora: float) -> Tuple[int, int, str, str]:
    valor_hora = max(valor_hora, 0.0)
    if valor_hora <= 0:
        tempo_maximo_min = 0
    else:
        tempo_maximo_min = int(round((row.mensalidade_eur / valor_hora) * 60))
    tempo_a_cortar_min = max(0, row.timing_media_minutos - tempo_maximo_min)
    return (
        tempo_maximo_min,
        tempo_a_cortar_min,
        _format_horas_minutos(tempo_maximo_min),
        _format_horas_minutos(tempo_a_cortar_min),
    )


def _render_excel_pretty(rows: List[ClienteRow], total_str: str, ano_sel: Optional[int], filtros: Dict[str, str], valor_hora: float = VALOR_HORA_EUR_DEFAULT) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Relacao Tecnicos"

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_title = PatternFill("solid", fgColor=_DARK_BLUE)
    fill_header = PatternFill("solid", fgColor=_GOLD)
    fill_alt = PatternFill("solid", fgColor=_LIGHT_GRAY)
    fill_alert = PatternFill("solid", fgColor=_ALERT_RED)
    fill_warn = PatternFill("solid", fgColor=_ALERT_AMBER)

    font_title = Font(bold=True, color=_WHITE, size=16)
    font_sub = Font(color=_WHITE, size=10)
    font_header = Font(bold=True, color="111827", size=11)
    font_total = Font(bold=True, color="111827", size=11)

    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # Título
    ws.merge_cells("A1:J1")
    ws["A1"] = "Relação Técnicos"
    ws["A1"].fill = fill_title
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:J2")
    ws["A2"] = f"PACACCOUNTING | Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}" + (f" | Ano: {ano_sel}" if ano_sel else "")
    ws["A2"].fill = fill_title
    ws["A2"].font = font_sub
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:J3")
    ws["A3"] = "Filtros: " + " | ".join([f"{k}={v}" for k, v in filtros.items() if v]) if any(filtros.values()) else "Filtros: (nenhum)"
    ws["A3"].fill = fill_title
    ws["A3"].font = font_sub
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

    ws.append([])

    headers = [
        "Nome",
        "NIF",
        "Técnico",
        "Técn. GRH",
        "Tipo contabilidade",
        "Periodicidade IVA",
        "Regime IVA",
        "Timing médio mensal",
        "Tempo máximo/mês",
        "Tempo a cortar",
    ]
    header_row = 5
    ws.append(headers)

    for col in range(1, 11):
        c = ws.cell(row=header_row, column=col)
        c.fill = fill_header
        c.font = font_header
        c.border = border
        c.alignment = align_center

    start_row = header_row + 1
    for i, r in enumerate(rows):
        rr = start_row + i
        ws.cell(rr, 1, r.nome).alignment = align_left
        ws.cell(rr, 2, r.nif).alignment = align_center
        ws.cell(rr, 3, r.tecnico).alignment = align_center
        ws.cell(rr, 4, r.tecnico_grh).alignment = align_center
        ws.cell(rr, 5, r.tipo_contabilidade).alignment = align_center
        ws.cell(rr, 6, r.periodicidade_iva).alignment = align_center
        ws.cell(rr, 7, r.regime_iva).alignment = align_center
        tempo_maximo_min, tempo_a_cortar_min, tempo_maximo_str, tempo_cortar_str = _calcular_tempo_limites(r, valor_hora)

        ws.cell(rr, 8, r.timing_media_str).alignment = align_right
        ws.cell(rr, 9, tempo_maximo_str).alignment = align_right
        ws.cell(rr, 10, tempo_cortar_str).alignment = align_right

        for ccol in range(1, 11):
            cell = ws.cell(rr, ccol)
            cell.border = border
            if i % 2 == 1:
                cell.fill = fill_alt

        # Destaques
        if "Sem timings" in r.qualidade:
            ws.cell(rr, 8).fill = fill_alert
        if not r.tecnico:
            ws.cell(rr, 3).fill = fill_warn
        if not r.tipo_contabilidade:
            ws.cell(rr, 5).fill = fill_warn
        if not r.periodicidade_iva:
            ws.cell(rr, 6).fill = fill_warn
        if not r.regime_iva:
            ws.cell(rr, 7).fill = fill_warn

    total_row = start_row + len(rows)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=7)
    ws.cell(total_row, 1, "TOTAL").font = font_total
    ws.cell(total_row, 1).alignment = align_right

    total_tempo_maximo = sum(_calcular_tempo_limites(r, valor_hora)[0] for r in rows)
    total_tempo_cortar = sum(_calcular_tempo_limites(r, valor_hora)[1] for r in rows)

    ws.cell(total_row, 8, total_str).font = font_total
    ws.cell(total_row, 8).alignment = align_right
    ws.cell(total_row, 9, _format_horas_minutos(total_tempo_maximo)).font = font_total
    ws.cell(total_row, 9).alignment = align_right
    ws.cell(total_row, 10, _format_horas_minutos(total_tempo_cortar)).font = font_total
    ws.cell(total_row, 10).alignment = align_right

    for col in range(1, 11):
        cell = ws.cell(total_row, col)
        cell.border = border
        cell.fill = fill_header

    widths = [40, 14, 18, 18, 20, 18, 18, 18, 18, 18]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = ws["A6"]
    ws.auto_filter.ref = f"A{header_row}:J{total_row}"

    try:
        tab = XLTable(displayName="RelacaoTecnicos", ref=f"A{header_row}:J{total_row}")
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
    except Exception:
        pass

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# =========================
# Export PDF bonito e estável (ReportLab)
# =========================

def _pdf_register_fonts() -> Tuple[str, str]:
    base_font = "Helvetica"
    bold_font = "Helvetica-Bold"
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        regular = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
        bold = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
        if os.path.exists(regular) and os.path.exists(bold):
            pdfmetrics.registerFont(TTFont("DejaVu", regular))
            pdfmetrics.registerFont(TTFont("DejaVuBold", bold))
            base_font = "DejaVu"
            bold_font = "DejaVuBold"
    except Exception:
        pass
    return base_font, bold_font


def _render_pdf_pretty(
    rows: List[ClienteRow],
    total_str: str,
    ano_sel: Optional[int],
    filtros: Dict[str, str],
    valor_hora: float = VALOR_HORA_EUR_DEFAULT,
) -> BytesIO:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Spacer, Paragraph, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail="Para exportar PDF, instala a biblioteca: pip install reportlab",
        ) from e

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="Relação Técnicos",
    )

    base_font, bold_font = _pdf_register_fonts()
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="TitlePAC",
            parent=styles["Title"],
            fontName=bold_font,
            fontSize=18,
            textColor=colors.HexColor("#061A44"),
        )
    )
    styles.add(
        ParagraphStyle(
            name="MetaPAC",
            parent=styles["Normal"],
            fontName=base_font,
            fontSize=9,
            textColor=colors.HexColor("#111827"),
            spaceAfter=2,
        )
    )

    header_bg = colors.HexColor("#D4AF37")

    def _footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont(base_font, 8)
        canvas.setFillColor(colors.grey)
        canvas.drawRightString(285 * mm, 8 * mm, f"Página {canvas.getPageNumber()}")
        canvas.restoreState()

    story: List[Any] = []
    story.append(Paragraph("Relação Técnicos", styles["TitlePAC"]))
    meta = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}" + (f" | Ano: {ano_sel}" if ano_sel else "")
    story.append(Paragraph(meta, styles["MetaPAC"]))
    filtros_visiveis = {k: v for k, v in filtros.items() if v and k != "valor_hora"}
    filt_txt = " | ".join([f"{k}={v}" for k, v in filtros_visiveis.items()]) if filtros_visiveis else "Sem filtros"
    story.append(Paragraph(f"Filtros: {filt_txt}", styles["MetaPAC"]))
    story.append(Spacer(1, 6))

    data: List[List[str]] = [[
        "Nome",
        "NIF",
        "Técnico",
        "Técn. GRH",
        "Tipo contab.",
        "Per. IVA",
        "Regime IVA",
        "Timing médio",
        "Tempo máximo/mês",
        "Tempo a cortar",
    ]]

    total_tempo_maximo = 0
    total_tempo_cortar = 0

    for r in rows:
        tempo_maximo_min, tempo_a_cortar_min, tempo_maximo_str, tempo_cortar_str = _calcular_tempo_limites(r, valor_hora)
        total_tempo_maximo += tempo_maximo_min
        total_tempo_cortar += tempo_a_cortar_min
        data.append([
            r.nome,
            r.nif or "—",
            r.tecnico or "—",
            r.tecnico_grh or "—",
            r.tipo_contabilidade or "—",
            r.periodicidade_iva or "—",
            r.regime_iva or "—",
            r.timing_media_str,
            tempo_maximo_str,
            tempo_cortar_str,
        ])

    data.append([
        "",
        "",
        "",
        "",
        "",
        "",
        "TOTAL",
        total_str,
        _format_horas_minutos(total_tempo_maximo),
        _format_horas_minutos(total_tempo_cortar),
    ])

    col_widths = [
        72 * mm,
        20 * mm,
        26 * mm,
        24 * mm,
        24 * mm,
        20 * mm,
        22 * mm,
        26 * mm,
        26 * mm,
        26 * mm,
    ]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    alt_bg = colors.HexColor("#F3F4F6")

    style_cmds: List[Tuple] = [
        ("FONTNAME", (0, 0), (-1, 0), bold_font),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ("BACKGROUND", (0, 0), (-1, 0), header_bg),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("ALIGN", (1, 1), (6, -2), "CENTER"),
        ("ALIGN", (7, 1), (-1, -2), "RIGHT"),
        ("FONTNAME", (0, 1), (-1, -2), base_font),
        ("FONTSIZE", (0, 1), (-1, -2), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#9CA3AF")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, -1), (-1, -1), bold_font),
        ("BACKGROUND", (0, -1), (-1, -1), header_bg),
        ("ALIGN", (6, -1), (6, -1), "RIGHT"),
        ("ALIGN", (7, -1), (-1, -1), "RIGHT"),
    ]

    for i in range(1, len(data) - 1):
        if i % 2 == 0:
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), alt_bg))

    for idx_row, r in enumerate(rows, start=1):
        if "Sem timings" in r.qualidade:
            style_cmds.append(("BACKGROUND", (7, idx_row), (7, idx_row), colors.HexColor("#FEE2E2")))

    table.setStyle(TableStyle(style_cmds))
    story.append(table)
    story.append(Spacer(1, 6))
    story.append(Paragraph("Nota: Timing médio mensal = (Σ meses / 12) + extra mensal.", styles["MetaPAC"]))

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    buffer.seek(0)
    return buffer


# =========================
# Rotas
# =========================

@router.get("/relacao-tecnicos", response_class=HTMLResponse)
async def pagina_relacao_tecnicos(request: Request):
    dados = _dataset(request)
    # Estes nomes estão alinhados com o teu relacao_tecnicos.html
    contexto = {
        "request": request,
        "linhas": dados["rows"],
        "filtros": dados["filtros"],
        "opcoes": dados["opcoes"],
        "sort_by": dados["sort_by"],
        "direction": dados["direction"],
        "sort_links": dados["sort_links"],
        "total_str": dados["total_str"],
        "contagem": dados["contagem"],
        "anos_disponiveis": dados["anos_disponiveis"],
        "ano_sel": dados["ano_sel"],
        "export_qs": dados["export_qs"],
        "export_qs_sem_tecnico": dados["export_qs_sem_tecnico"],
        "tecnico_blocks": dados["tecnico_blocks"],
        "tecnicos_lista": dados["tecnicos_lista"],
    }
    return templates.TemplateResponse("relacao_tecnicos.html", contexto)


def _resolver_valor_hora(request: Request, valor_query: str | None) -> float:
    if valor_query:
        try:
            return max(0.0, float(str(valor_query).replace(",", ".")))
        except Exception:
            pass
    q_val = request.query_params.get("valor_hora")
    if q_val:
        try:
            return max(0.0, float(str(q_val).replace(",", ".")))
        except Exception:
            pass
    return VALOR_HORA_EUR_DEFAULT


@router.get("/relacao-tecnicos/download/excel")
async def exportar_excel_relacao_tecnico(request: Request, tecnico: str | None = None, valor_hora: str | None = None):
    valor_hora_eur = _resolver_valor_hora(request, valor_hora)
    linhas, total_str, filtros_export, ano_sel, tecnico_display = _prepare_tecnico_export(request, tecnico, valor_hora_eur)
    buffer = _render_excel_pretty(linhas, total_str, ano_sel, filtros_export, valor_hora_eur)
    data_stamp = datetime.now().strftime("%Y-%m-%d")
    filename = f"relacao_tecnicos_{_slugify_tecnico_filename(tecnico_display)}_{data_stamp}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/relacao-tecnicos/download/pdf")
async def exportar_pdf_relacao_tecnico(request: Request, tecnico: str | None = None, valor_hora: str | None = None):
    valor_hora_eur = _resolver_valor_hora(request, valor_hora)
    linhas, total_str, filtros_export, ano_sel, tecnico_display = _prepare_tecnico_export(request, tecnico, valor_hora_eur)
    buffer = _render_pdf_pretty(linhas, total_str, ano_sel, filtros_export, valor_hora_eur)
    data_stamp = datetime.now().strftime("%Y-%m-%d")
    filename = f"relacao_tecnicos_{_slugify_tecnico_filename(tecnico_display)}_{data_stamp}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/relacao-tecnicos/export/excel")
async def exportar_excel_relacao_tecnicos(request: Request):
    dados = _dataset(request)
    valor_hora_eur = _resolver_valor_hora(request, None)
    filtros_export = dict(dados["filtros"])
    filtros_export["valor_hora"] = f"{valor_hora_eur:.2f}"
    buffer = _render_excel_pretty(dados["rows"], dados["total_str"], dados["ano_sel"], filtros_export, valor_hora_eur)
    filename = "relacao_tecnicos.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/relacao-tecnicos/export/pdf")
async def exportar_pdf_relacao_tecnicos(request: Request):
    dados = _dataset(request)
    valor_hora_eur = _resolver_valor_hora(request, None)
    filtros_export = dict(dados["filtros"])
    filtros_export["valor_hora"] = f"{valor_hora_eur:.2f}"
    buffer = _render_pdf_pretty(dados["rows"], dados["total_str"], dados["ano_sel"], filtros_export, valor_hora_eur)
    filename = "relacao_tecnicos.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
