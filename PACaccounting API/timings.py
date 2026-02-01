from fastapi import APIRouter, Request, UploadFile, File, Form, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates

from io import BytesIO
import re
import json
import os
import shutil
from datetime import datetime
import unicodedata
from typing import List, Dict, Optional, Any, Tuple
from urllib.parse import urlencode

import openpyxl  # pip install openpyxl

# fallback: se timings_dados.json estiver vazio, vamos buscar aos dados gerais
from dados import estado, guardar_dados

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# Ficheiro próprio de timings (independente de dados.json)
TIMINGS_FILE = os.path.join(BASE_DIR, "timings_dados.json")
TIMINGS_IMPORT_REPORT = os.path.join(BASE_DIR, "timings_import_report.json")

print("[TIMINGS][PATH] cwd=", os.getcwd())
print("[TIMINGS][PATH] __file__ dir=", BASE_DIR)
print(
    "[TIMINGS][PATH] TIMINGS_FILE=",
    TIMINGS_FILE,
    "exists=",
    os.path.exists(TIMINGS_FILE),
)

router = APIRouter()
templates = Jinja2Templates(directory="templates")


# ========= HELPERS DE NOME / CONFIG =========

def _normalize_nome(s: str) -> str:
    """
    Normaliza um nome:
      - strip
      - maiúsculas
      - remove acentos
      - comprime espaços múltiplos num só
    """
    if s is None:
        return ""
    s = s.strip().upper()
    # remover acentos
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    # comprimir espaços
    s = re.sub(r"\s+", " ", s)
    return s


LEGAL_SUFFIX_PATTERN = re.compile(
    r"\b(?:lda|ltda|unipessoal|sociedade|por|quotas?|sa|s\s+a)\b",
    re.IGNORECASE,
)


def _norm_empresa_forte(s: Optional[str]) -> str:
    """Normaliza fortemente nomes de empresas para facilitar matching."""
    if not s:
        return ""

    txt = unicodedata.normalize("NFD", str(s).strip())
    txt = "".join(ch for ch in txt if unicodedata.category(ch) != "Mn")
    txt = txt.casefold()
    if not txt:
        return ""

    txt = re.sub(r"[\.,;:\-_/()\[\]{}&'\"+]", " ", txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    if not txt:
        return ""

    original_txt = txt
    txt = LEGAL_SUFFIX_PATTERN.sub(" ", txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt if txt else original_txt


def _normalize_nome_empresa(s: str) -> str:
    """Compat: manter assinatura antiga."""
    return _norm_empresa_forte(s)


def _normalize_header(value: Optional[str]) -> str:
    """Normaliza cabeçalhos de Excel para comparação case-insensitive sem acentos."""
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = s.lower()
    s = s.replace("/", " ")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return s.strip()


def _norm_nome_forte(s: Optional[str]) -> str:
    """Normaliza fortemente nomes: maiúsculas, sem acentos, sem partículas comuns."""
    if not s:
        return ""
    s = unicodedata.normalize("NFD", str(s).strip())
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = s.upper()
    tokens = re.split(r"\s+", s)
    stopwords = {"DE", "DA", "DO", "DOS", "DAS", "E"}
    tokens = [t for t in tokens if t and t not in stopwords]
    return " ".join(tokens)


def _build_aliases_canonicos() -> Dict[str, str]:
    """Constrói o mapa de aliases fortemente normalizados -> nome canónico."""
    alias_map: Dict[str, str] = {}

    def registrar(canonico: str, *variantes: str) -> None:
        nomes = (canonico, *variantes)
        for nome in nomes:
            norm = _norm_nome_forte(nome)
            if norm:
                alias_map[norm] = canonico

    registrar(
        "Pedro Fernandes",
        "Pedro Miguel da Silva Fernandes",
        "Pedro Fernandes",
    )
    registrar(
        "Ana Rodrigues",
        "Ana Catarina Lourenço Rodrigues",
        "Ana Catarina Lorenco Rodrigues",
        "Ana Rodrigues",
    )
    registrar(
        "Daniela Fernandes",
        "Marta Daniela Francisco Fernandes",
        "Daniela Francisco Fernandes",
    )
    registrar(
        "Celine Santos",
        "Celine",
        "Celine Rodrigues dos Santos",
        "Celine Santos",
        "ANTONIO CANDIDO GONÇALVES DIAS",
        "ANTONIO CANDIDO GONCALVES DIAS",
    )
    registrar(
        "M Albertina Alves",
        "Maria Albertina Pereira Alves",
        "M Albertina Alves",
    )
    registrar(
        "M Luzia Moreira",
        "Luzia Maria Gonçalves Moreira",
        "Luzia Maria Goncalves Moreira",
        "M Luzia Moreira",
    )
    registrar(
        "João Pedro Alves",
        "João Pedro Gonçalves Alves",
        "Joao Pedro Goncalves Alves",
        "João Pedro Alves",
        "Joao Pedro Alves",
    )

    return alias_map


ALIASES_CANONICOS = _build_aliases_canonicos()
ARMANDO_NORMS = {
    _norm_nome_forte("Armando Palhão Dias"),
    _norm_nome_forte("Armando Palhao Dias"),
    _norm_nome_forte("Armando Dias"),
}


def _resolver_tecnico(nome_excel: Optional[str]) -> tuple[str, Optional[str]]:
    """Classifica um nome proveniente do Excel."""
    if not nome_excel:
        return "desconhecido", None

    norm_forte = _norm_nome_forte(nome_excel)
    if not norm_forte:
        return "desconhecido", None

    if norm_forte in ARMANDO_NORMS:
        return "armando", None

    canonico = ALIASES_CANONICOS.get(norm_forte)
    if canonico:
        return "canonico", canonico

    return "desconhecido", None


def _canonical_tecnico_nome(s: Optional[str]) -> str:
    """Converte qualquer nome para a forma canónica conhecida."""
    if not s:
        return "Sem técnico"

    tipo, canonico = _resolver_tecnico(s)
    if tipo == "canonico" and canonico:
        return canonico
    if tipo == "armando":
        return "Sem técnico"
    s = str(s).strip()
    return s or "Sem técnico"

EMPRESA_HEADER_NAMES = {
    "empresa",
    "cliente",
    "cliente empresa",
    "empresa cliente",
    "designacao",
    "designacao social",
    "nome cliente",
    "cliente nome",
}

TECNICO_HEADER_NAMES = {
    "tecnico",
    "tecnico responsavel",
    "responsavel",
    "responsavel tecnico",
    "colaborador",
    "tecnica",
}

TEMPO_HEADER_NAMES = {
    "tempo",
    "horas",
    "horas totais",
    "tempo m",
    "tempo min",
    "tempo minutos",
    "minutos",
    "total minutos",
    "duracao",
    "duracao minutos",
    "duracao m",
}

MESES_LABELS = [
    (1, "Jan"),
    (2, "Fev"),
    (3, "Mar"),
    (4, "Abr"),
    (5, "Mai"),
    (6, "Jun"),
    (7, "Jul"),
    (8, "Ago"),
    (9, "Set"),
    (10, "Out"),
    (11, "Nov"),
    (12, "Dez"),
]


def _is_linha_total(texto: Optional[str]) -> bool:
    if not texto:
        return False
    norm = _normalize_header(texto)
    if not norm:
        return False
    return any(kw in norm for kw in {"total", "subtotal", "grand total", "soma", "sum"})

# Estrutura em memória (formato novo):
# {
#   "2025": {
#       "Empresa X": {
#           "meses": {1: 120, 2: 60, ...},
#           "extra_mensal": 15,
#           "apagado": False,
#           "por_tecnico": {...}
#       },
#       ...
#   },
#   ...
# }
timings_dados: dict = {}
_PRECISA_BACKUP_TIMINGS = False
_PRECISA_REGRAVAR_TIMINGS = False


# ========= HELPERS DE TEMPO =========

def _format_minutos(minutos: int) -> str:
    """Converte minutos em 'XhYYm'."""
    if minutos is None:
        minutos = 0
    minutos = int(minutos)
    horas = minutos // 60
    mins = minutos % 60
    return f"{horas}h{mins:02d}m"


def _parse_duracao_para_minutos(valor) -> int:
    """Interpreta horas/minutos em vários formatos e devolve minutos inteiros."""
    if valor is None:
        return 0

    s = str(valor).strip()
    if not s:
        return 0

    s_lower = s.lower()
    match_h = re.search(r"([0-9]+(?:[.,][0-9]+)?)\s*h", s_lower)
    match_m = re.search(r"([0-9]+)\s*m", s_lower)

    horas_float = 0.0
    minutos_int = 0

    if match_h:
        try:
            horas_float = float(match_h.group(1).replace(",", "."))
        except ValueError:
            horas_float = 0.0

    if match_m:
        try:
            minutos_int = int(match_m.group(1))
        except ValueError:
            minutos_int = 0

    if match_h or match_m:
        total = int(round(horas_float * 60)) + max(0, minutos_int)
        return max(0, total)

    normalizado = s.replace(",", ".")

    if "." in normalizado:
        try:
            horas = float(normalizado)
        except ValueError:
            return 0
        if abs(horas) <= 24:
            return max(0, int(round(horas * 60)))
        return max(0, int(round(horas)))

    try:
        minutos = int(round(float(normalizado)))
    except ValueError:
        return 0
    return max(0, minutos)


def _parse_tempo_para_minutos(valor) -> int:
    """Compatibilidade retroativa para código legado."""
    return _parse_duracao_para_minutos(valor)


# ========= HELPERS DE ESTATÍSTICA =========

def _extra_mes(rec: Optional[dict]) -> int:
    """Obtém o extra mensal médio (minutos) para um registo normalizado."""
    if not isinstance(rec, dict):
        return 0
    valor = rec.get("extra_mensal", 0)
    if isinstance(valor, (int, float)):
        return int(valor)
    try:
        return _parse_duracao_para_minutos(valor)
    except Exception:
        return 0


def _total_minutos_timings(data: dict) -> int:
    """Soma todos os minutos (meses + extra_mensal*12) de uma estrutura timings."""
    if not isinstance(data, dict):
        return 0

    total = 0
    for ano_val in data.values():
        if not isinstance(ano_val, dict):
            continue
        for empresa in ano_val.values():
            if not isinstance(empresa, dict):
                continue
            meses = empresa.get("meses", {})
            if isinstance(meses, dict):
                for valor in meses.values():
                    total += _parse_duracao_para_minutos(valor)
            total += _extra_mes(empresa) * 12
    return total


def _normalizar_dados_timings_brutos(data: dict) -> dict:
    """Produz uma cópia normalizada (minutos inteiros, estruturas coerentes)."""
    global _PRECISA_BACKUP_TIMINGS, _PRECISA_REGRAVAR_TIMINGS
    resultado: Dict[str, Dict[str, dict]] = {}
    if not isinstance(data, dict):
        return resultado

    houve_migracao_por_tecnico = False

    for ano_key, ano_val in data.items():
        if not isinstance(ano_val, dict):
            continue
        ano_str = str(ano_key)
        ano_dict = resultado.setdefault(ano_str, {})

        for empresa_nome, rec in ano_val.items():
            if not isinstance(rec, dict):
                continue

            empresa_display = str(empresa_nome)
            rec_norm: Dict[str, Any] = {}

            meses_raw = rec.get("meses", {})
            meses_norm: Dict[int, int] = {}
            if isinstance(meses_raw, dict):
                for mes_key, valor in meses_raw.items():
                    try:
                        mes_int = int(mes_key)
                    except (TypeError, ValueError):
                        continue
                    minutos_val = _parse_duracao_para_minutos(valor)
                    if minutos_val > 0:
                        meses_norm[mes_int] = minutos_val
            rec_norm["meses"] = meses_norm

            extra_mensal_raw = rec.get("extra_mensal")
            if extra_mensal_raw is not None:
                extra_mensal = _parse_duracao_para_minutos(extra_mensal_raw)
            else:
                legacy_extra = rec.get("extra")
                if legacy_extra is not None:
                    legacy_total = _parse_duracao_para_minutos(legacy_extra)
                    if legacy_total > 0:
                        extra_mensal = int(round(legacy_total / 12))
                    else:
                        extra_mensal = 0
                    if extra_mensal > 0:
                        _PRECISA_BACKUP_TIMINGS = True
                else:
                    extra_mensal = 0

            rec_norm["extra_mensal"] = max(0, int(extra_mensal))
            rec_norm["apagado"] = bool(rec.get("apagado", False))

            por_tecnico_raw = rec.get("por_tecnico")
            por_tecnico_norm: Dict[str, Dict[str, int]] = {}
            if isinstance(por_tecnico_raw, dict):
                for tec_nome, meses_tecnico in por_tecnico_raw.items():
                    if not isinstance(meses_tecnico, dict):
                        continue

                    meses_tecnico_norm: Dict[str, int] = {}
                    for mes_key, valor in meses_tecnico.items():
                        try:
                            mes_int = int(mes_key)
                        except (TypeError, ValueError):
                            continue
                        minutos_val = _parse_duracao_para_minutos(valor)
                        if minutos_val > 0:
                            meses_tecnico_norm[str(mes_int)] = minutos_val

                    if not meses_tecnico_norm:
                        continue

                    tec_nome_str = str(tec_nome).strip() if tec_nome is not None else ""
                    tecnico_canonico = _canonical_tecnico_nome(tec_nome_str)
                    if not tecnico_canonico:
                        tecnico_canonico = _canonical_tecnico_nome(None)

                    destino = por_tecnico_norm.setdefault(tecnico_canonico, {})
                    for mes_key_norm, minutos_val in meses_tecnico_norm.items():
                        destino[mes_key_norm] = int(destino.get(mes_key_norm, 0) or 0) + minutos_val

                    if tecnico_canonico != tec_nome_str:
                        houve_migracao_por_tecnico = True
                        _PRECISA_BACKUP_TIMINGS = True
            rec_norm["por_tecnico"] = por_tecnico_norm

            ano_dict[empresa_display] = rec_norm

    if houve_migracao_por_tecnico:
        _PRECISA_REGRAVAR_TIMINGS = True

    return resultado


# ========= HELPERS DE PERSISTÊNCIA =========

def _guardar_timings_para_ficheiro() -> bool:
    """Guarda todo o dicionário timings_dados em timings_dados.json."""
    novo_total = _total_minutos_timings(timings_dados)

    total_anterior = None
    if os.path.exists(TIMINGS_FILE):
        try:
            with open(TIMINGS_FILE, "r", encoding="utf-8") as f:
                antigo = json.load(f)
            total_anterior = _total_minutos_timings(antigo)
        except Exception:
            total_anterior = None

    if total_anterior and total_anterior > 0:
        if novo_total < total_anterior * 0.5:
            print(
                "[TIMINGS] WARNING: tentativa de gravação rejeitada (redução superior a 50% nos minutos totais)."
            )
            return False

    tmp = TIMINGS_FILE + ".tmp"
    destino = os.path.abspath(TIMINGS_FILE)
    try:
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(timings_dados, f, ensure_ascii=False, indent=2)
        os.replace(tmp, TIMINGS_FILE)
        print(f"[TIMINGS] Guardado em disco: {destino}")
        return True
    except Exception as exc:
        print(f"[TIMINGS] ERRO ao escrever timings_dados.json: {exc}")
        if os.path.exists(tmp):
            try:
                os.remove(tmp)
            except Exception:
                pass
        return False


def _migrar_de_legacy_dict(legacy_timings: dict, legacy_extras: dict) -> dict:
    """
    Converte formato antigo:
      legacy_timings = {
          "2025": {
              "empresas": {
                  "Empresa X": { "1": 120, "2": 60, ... },
                  ...
              }
          }
      }
      legacy_extras = {
          "2025": {
              "empresas": {
                  "Empresa X": 45,
                  ...
              }
          }
      }
    para o formato novo:
            {
                "2025": {
                        "Empresa X": {
                                "meses": {1: 120, 2: 60, ...},
                                "extra_mensal": 45,
                                "apagado": False
                        },
                        ...
                }
            }
    """
    global _PRECISA_BACKUP_TIMINGS
    new: dict = {}

    if not isinstance(legacy_timings, dict):
        return {}

    for ano_key, ano_val in legacy_timings.items():
        if not isinstance(ano_val, dict):
            continue
        empresas = ano_val.get("empresas", {})
        if not isinstance(empresas, dict):
            continue

        ano_str = str(ano_key)
        ano_new = new.setdefault(ano_str, {})

        extras_ano_empresas = {}
        if isinstance(legacy_extras, dict):
            extras_ano = legacy_extras.get(ano_str) or legacy_extras.get(ano_key) or {}
            if isinstance(extras_ano, dict):
                extras_ano_empresas = extras_ano.get("empresas", {})
                if not isinstance(extras_ano_empresas, dict):
                    extras_ano_empresas = {}

        for emp, meses_dict in empresas.items():
            if not isinstance(meses_dict, dict):
                continue

            meses_int = {}
            for mes_key, v in meses_dict.items():
                try:
                    mes_int = int(mes_key)
                except (TypeError, ValueError):
                    continue
                mins_int = _parse_duracao_para_minutos(v)
                meses_int[mes_int] = mins_int

            extra_mensal = 0
            if extras_ano_empresas:
                extra_val = extras_ano_empresas.get(emp)
                if extra_val is not None:
                    extra_total = _parse_duracao_para_minutos(extra_val)
                    if extra_total > 0:
                        extra_mensal = int(round(extra_total / 12))
                        _PRECISA_BACKUP_TIMINGS = True

            ano_new[emp] = {
                "meses": meses_int,
                "extra_mensal": extra_mensal,
                "apagado": False,
            }

    return new


def _persistir_timings() -> None:
    """Grava timings_dados e sincroniza com dados.json."""
    global _PRECISA_BACKUP_TIMINGS

    if _PRECISA_BACKUP_TIMINGS and os.path.exists(TIMINGS_FILE):
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        backup_path = f"{TIMINGS_FILE}.bak.{timestamp}"
        try:
            shutil.copyfile(TIMINGS_FILE, backup_path)
            print(f"[TIMINGS] Backup criado: {os.path.abspath(backup_path)}")
            _PRECISA_BACKUP_TIMINGS = False
        except Exception as exc:
            print(f"[TIMINGS] WARNING: falha ao criar backup antes da migração: {exc}")

    guardado = _guardar_timings_para_ficheiro()
    if not guardado:
        return

    try:
        estado["timings_dados"] = timings_dados
        guardar_dados()
    except Exception as exc:
        print(f"[TIMINGS] ERRO ao atualizar dados.json: {exc}")


def _persistir_timings_se_preciso() -> None:
    """Persistência condicional após migrações em memória."""
    global _PRECISA_REGRAVAR_TIMINGS
    if not _PRECISA_REGRAVAR_TIMINGS:
        return
    _persistir_timings()
    _PRECISA_REGRAVAR_TIMINGS = False


def _carregar_timings_de_ficheiro() -> None:
    """
    Lê timings_dados.json e aceita:
      - formato novo (direto)
      - formato antigo (com 'timings'/'timings_extra')
    Se não encontrar nada válido, faz fallback a estado["timings"] / ["timings_extra"] de dados.py.
    """
    global timings_dados

    data = None

    if os.path.exists(TIMINGS_FILE):
        try:
            with open(TIMINGS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            data = None

    # 1) Se o ficheiro já estiver no formato novo (dict de anos, sem 'timings'/'timings_extra')
    if isinstance(data, dict) and "timings" not in data and "timings_extra" not in data:
        timings_dados = _normalizar_dados_timings_brutos(data)
        estado["timings_dados"] = timings_dados
        _persistir_timings_se_preciso()
        return

    # 2) Se estiver no formato antigo (timings/timings_extra) dentro do ficheiro próprio
    if isinstance(data, dict) and ("timings" in data or "timings_extra" in data):
        legacy_timings = data.get("timings", {})
        legacy_extras = data.get("timings_extra", {})
        timings_dados = _migrar_de_legacy_dict(legacy_timings, legacy_extras)
        estado["timings_dados"] = timings_dados
        _persistir_timings_se_preciso()
        return

    # 3) Fallback: tentar ir buscar diretamente a estado["timings"] / ["timings_extra"]
    try:
        estado_timings = estado.get("timings_dados")
    except Exception:
        estado_timings = None

    if isinstance(estado_timings, dict) and estado_timings:
        timings_dados = _normalizar_dados_timings_brutos(estado_timings)
        estado["timings_dados"] = timings_dados
        _persistir_timings_se_preciso()
        return

    try:
        legacy_timings = estado.get("timings", {})
        legacy_extras = estado.get("timings_extra", {})
    except Exception:
        legacy_timings = {}
        legacy_extras = {}

    if legacy_timings or legacy_extras:
        timings_dados = _migrar_de_legacy_dict(legacy_timings, legacy_extras)
    else:
        timings_dados = {}

    estado["timings_dados"] = timings_dados
    _persistir_timings_se_preciso()


_carregar_timings_de_ficheiro()


def _migrar_timings_para_minutos() -> Tuple[bool, str]:
    """Executa migração manual para converter horas decimais em minutos inteiros."""
    if not os.path.exists(TIMINGS_FILE):
        return False, "timings_dados.json não encontrado."

    try:
        with open(TIMINGS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as exc:
        return False, f"erro ao ler ficheiro: {exc}"

    if isinstance(data, dict) and ("timings" in data or "timings_extra" in data):
        normalizado = _migrar_de_legacy_dict(data.get("timings", {}), data.get("timings_extra", {}))
    else:
        normalizado = _normalizar_dados_timings_brutos(data)

    novo_total = _total_minutos_timings(normalizado)
    antigo_total = _total_minutos_timings(data)

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = f"{TIMINGS_FILE}.bak.{timestamp}"
    try:
        shutil.copyfile(TIMINGS_FILE, backup_path)
    except Exception as exc:
        return False, f"não foi possível criar backup: {exc}"

    global timings_dados
    timings_dados = normalizado

    return True, f"totais minutos antigos={antigo_total}, novos={novo_total}, backup={backup_path}"


def _mapear_clientes_por_nome():
    """
    Cria:
      - um dicionário chave_normalizada -> cliente
      - uma lista [(chave_normalizada, cliente), ...]
    a partir de estado["clientes"], usando _norm_empresa_forte,
    para podermos fazer matches exatos e, em último caso, por inclusão.
    """
    mapa = {}
    lista_norm = []
    try:
        clientes = estado.get("clientes", [])
    except Exception:
        clientes = []
    if not isinstance(clientes, list):
        return {}, []

    for cli in clientes:
        if not isinstance(cli, dict):
            continue
        nome = (
            cli.get("nome")
            or cli.get("cliente")
            or cli.get("empresa")
            or cli.get("designacao")
        )
        if not nome:
            continue
        norm = _norm_empresa_forte(str(nome))
        if not norm:
            continue

        mapa[norm] = cli
        lista_norm.append((norm, cli))

    return mapa, lista_norm


def _obter_tecnico_cliente(cli: dict) -> Optional[str]:
    """
    Tenta obter o nome do técnico associado a um cliente.
    Procura por ordem nos campos: 'tecnico', 'carteira', 'responsavel'.
    """
    if not isinstance(cli, dict):
        return None
    for campo in ("tecnico", "carteira", "responsavel"):
        val = cli.get(campo)
        if val:
            s = str(val).strip()
            if s:
                return s
    return None


def _tecnico_do_cliente(cli: Optional[dict]) -> str:
    """Devolve o técnico canónico definido no cliente (ou 'Sem técnico')."""
    if not isinstance(cli, dict):
        return _canonical_tecnico_nome(None)

    tecnico_raw = cli.get("tecnico")
    if not tecnico_raw:
        return _canonical_tecnico_nome(None)

    return _canonical_tecnico_nome(tecnico_raw)


def _tecnico_inferido_empresa(empresa_norm: str) -> Optional[str]:
    """Determina o técnico (canónico) associado a uma empresa via estado['clientes']."""
    if not empresa_norm:
        return None

    alvo_norm = _norm_empresa_forte(empresa_norm)
    if not alvo_norm:
        return None

    try:
        clientes = estado.get("clientes", [])
    except Exception:
        clientes = []

    if not isinstance(clientes, list):
        return None

    for cli in clientes:
        if not isinstance(cli, dict):
            continue

        nome_cli = (
            cli.get("nome")
            or cli.get("cliente")
            or cli.get("empresa")
            or cli.get("designacao")
        )
        if not nome_cli:
            continue

        nome_norm = _norm_empresa_forte(nome_cli)
        if nome_norm != alvo_norm:
            continue

        tecnico_raw = cli.get("tecnico") or cli.get("carteira")
        if not tecnico_raw:
            return None

        tipo, canonico = _resolver_tecnico(tecnico_raw)
        if tipo == "canonico" and canonico:
            return canonico

        tecnico_str = str(tecnico_raw).strip()
        return tecnico_str or None

    return None


# ========= LÓGICA DE NEGÓCIO =========

def _obter_ano_dict(ano: int) -> dict:
    """Devolve (e cria se não existir) o dicionário de empresas desse ano."""
    return timings_dados.setdefault(str(ano), {})


def _adicionar_tempo_empresa(
    ano: int,
    mes: int,
    empresa: str,
    minutos: int,
    tecnico: Optional[str] = None,
) -> None:
    """
    Soma minutos a uma empresa num dado ano/mês em timings_dados (em memória)
    e grava imediatamente no ficheiro.
    """
    if minutos <= 0:
        return

    ano_dict = _obter_ano_dict(ano)
    emp_dict = ano_dict.setdefault(
        empresa,
        {"meses": {}, "extra_mensal": 0, "apagado": False, "por_tecnico": {}},
    )
    # se tinha sido apagado e volta a ter tempos, reativa
    emp_dict["apagado"] = False
    meses = emp_dict.setdefault("meses", {})
    atual = int(meses.get(mes, 0) or 0)
    meses[mes] = atual + minutos

    if tecnico:
        tecnico_norm = _canonical_tecnico_nome(tecnico)
        tec_dict = emp_dict.setdefault("por_tecnico", {})
        tec_registos = tec_dict.setdefault(tecnico_norm, {})
        mes_key = str(mes)
        atual_tec = int(tec_registos.get(mes_key, 0) or 0)
        tec_registos[mes_key] = atual_tec + minutos


def _encontrar_empresa_existente_por_norm(ano_dict: dict, empresa_norm: str) -> Optional[str]:
    """Procura chave de empresa existente cujo nome normalizado coincide com empresa_norm."""
    for nome in ano_dict.keys():
        if _norm_empresa_forte(nome) == empresa_norm:
            return nome
    return None


def _processar_sheet_workload(
    sheet,
    invalidos: Optional[Dict[str, int]] = None,
    ignorados_por_empresa: Optional[Dict[str, int]] = None,
    resumos_ignorados_por_empresa: Optional[Dict[str, int]] = None,
) -> tuple[List[tuple[str, str, Optional[str], int]], Dict[str, int]]:
    """
    Lê uma folha no formato legacy (empresa + técnicos indentados).
    Devolve uma lista de registos detalhados (empresa, tipo, canónico, minutos)
    e um dicionário com totais de fallback (quando só existe linha-resumo).
    """
    registos: List[tuple[str, str, Optional[str], int]] = []
    totais_fallback: Dict[str, int] = {}

    empresa_atual: Optional[str] = None
    resumo_atual_min = 0
    registos_empresa: List[tuple[str, str, Optional[str], int]] = []

    def finalizar_empresa() -> None:
        nonlocal empresa_atual, resumo_atual_min, registos_empresa

        if not empresa_atual:
            return

        if registos_empresa:
            if resumo_atual_min > 0:
                if ignorados_por_empresa is not None:
                    ignorados_por_empresa[empresa_atual] = (
                        ignorados_por_empresa.get(empresa_atual, 0) + resumo_atual_min
                    )
                if resumos_ignorados_por_empresa is not None:
                    resumos_ignorados_por_empresa[empresa_atual] = (
                        resumos_ignorados_por_empresa.get(empresa_atual, 0) + resumo_atual_min
                    )
            registos.extend(registos_empresa)
        elif resumo_atual_min > 0:
            totais_fallback[empresa_atual] = (
                totais_fallback.get(empresa_atual, 0) + resumo_atual_min
            )

        empresa_atual = None
        resumo_atual_min = 0
        registos_empresa = []

    for row in sheet.iter_rows(values_only=True):
        if not row:
            continue

        col1 = row[0]
        col2 = row[1] if len(row) > 1 else None

        if col1 is None:
            continue

        texto = str(col1)
        texto_strip = texto.strip()

        if not texto_strip:
            continue
        if texto_strip.upper() in {"EMPRESA", "TOTAL"} or _is_linha_total(texto_strip):
            continue
        if "MAPA DE TEMPO TRABALHADO" in texto_strip.upper():
            continue

        if texto.startswith("    "):
            if not empresa_atual:
                continue

            minutos = _parse_tempo_para_minutos(col2)
            if minutos <= 0:
                continue

            tipo, canonico = _resolver_tecnico(texto_strip)
            if tipo == "desconhecido":
                if invalidos is not None:
                    chave = texto_strip
                    invalidos[chave] = invalidos.get(chave, 0) + minutos
                if ignorados_por_empresa is not None:
                    ignorados_por_empresa[empresa_atual] = (
                        ignorados_por_empresa.get(empresa_atual, 0) + minutos
                    )
                continue

            registos_empresa.append((empresa_atual, tipo, canonico, minutos))
        else:
            finalizar_empresa()
            empresa_atual = texto_strip
            resumo_atual_min = _parse_tempo_para_minutos(col2)

    finalizar_empresa()

    return registos, totais_fallback


def _processar_sheet_colunas(
    sheet,
    invalidos: Optional[Dict[str, int]] = None,
    ignorados_por_empresa: Optional[Dict[str, int]] = None,
    resumos_ignorados_por_empresa: Optional[Dict[str, int]] = None,
) -> Optional[tuple[List[tuple[str, str, Optional[str], int]], Dict[str, int]]]:
    """Tenta ler folha em formato tabular (empresa/cliente, técnico, tempo)."""
    linhas = list(sheet.iter_rows(values_only=True))
    idx_empresa = idx_tecnico = idx_tempo = None
    header_idx = None

    for i, row in enumerate(linhas):
        if not row:
            continue
        valores_norm = [_normalize_header(c) for c in row]
        if not any(valores_norm):
            continue

        for idx, nome in enumerate(valores_norm):
            if not nome:
                continue
            if idx_empresa is None and nome in EMPRESA_HEADER_NAMES:
                idx_empresa = idx
            if idx_tecnico is None and nome in TECNICO_HEADER_NAMES:
                idx_tecnico = idx
            if idx_tempo is None and nome in TEMPO_HEADER_NAMES:
                idx_tempo = idx

        if idx_empresa is not None and idx_tecnico is not None and idx_tempo is not None:
            header_idx = i
            break

    if header_idx is None:
        return None

    registos_por_empresa: Dict[str, List[tuple[str, Optional[str], int]]] = {}
    resumos_por_empresa: Dict[str, int] = {}

    for row in linhas[header_idx + 1 :]:
        if not row:
            continue

        if idx_empresa >= len(row) or idx_tecnico >= len(row) or idx_tempo >= len(row):
            continue

        empresa_val = row[idx_empresa]
        tempo_val = row[idx_tempo]

        if empresa_val is None:
            continue

        empresa = str(empresa_val).strip()
        if not empresa:
            continue
        if _is_linha_total(empresa):
            continue

        minutos = _parse_tempo_para_minutos(tempo_val)
        if minutos <= 0:
            continue

        tecnico_val = row[idx_tecnico] if idx_tecnico is not None else None
        tecnico = str(tecnico_val).strip() if tecnico_val is not None else ""

        if not tecnico:
            resumos_por_empresa[empresa] = resumos_por_empresa.get(empresa, 0) + minutos
            continue

        tipo, canonico = _resolver_tecnico(tecnico)
        if tipo == "desconhecido":
            if not _is_linha_total(tecnico):
                if invalidos is not None:
                    chave = tecnico
                    invalidos[chave] = invalidos.get(chave, 0) + minutos
                if ignorados_por_empresa is not None:
                    ignorados_por_empresa[empresa] = ignorados_por_empresa.get(empresa, 0) + minutos
            continue

        registos = registos_por_empresa.setdefault(empresa, [])
        registos.append((tipo, canonico, minutos))

    resultados: List[tuple[str, str, Optional[str], int]] = []
    totais_fallback: Dict[str, int] = {}

    for empresa, detalhes in registos_por_empresa.items():
        tem_detalhe = any(tipo in {"canonico", "armando"} for tipo, _, _ in detalhes)

        if tem_detalhe:
            resumo_min = resumos_por_empresa.pop(empresa, 0)
            if resumo_min > 0:
                if ignorados_por_empresa is not None:
                    ignorados_por_empresa[empresa] = ignorados_por_empresa.get(empresa, 0) + resumo_min
                if resumos_ignorados_por_empresa is not None:
                    resumos_ignorados_por_empresa[empresa] = (
                        resumos_ignorados_por_empresa.get(empresa, 0) + resumo_min
                    )
            for tipo, canonico, minutos in detalhes:
                if tipo in {"canonico", "armando"} and minutos > 0:
                    resultados.append((empresa, tipo, canonico, minutos))
        else:
            total_resumo = resumos_por_empresa.pop(empresa, 0)
            total = total_resumo + sum(minutos for _, _, minutos in detalhes)
            if total > 0:
                totais_fallback[empresa] = totais_fallback.get(empresa, 0) + total

    for empresa, minutos in resumos_por_empresa.items():
        if minutos > 0:
            totais_fallback[empresa] = totais_fallback.get(empresa, 0) + minutos

    return resultados, totais_fallback


def _importar_excel_timings(
    conteudo: bytes,
    ano: int,
    mes: int,
) -> tuple[List[Dict[str, Any]], Dict[str, int], Dict[str, int], Dict[str, int]]:
    """
    Importa um Excel de workload e agrega registos normalizados para posterior deduplicação.
    """
    wb = openpyxl.load_workbook(BytesIO(conteudo), data_only=True)
    folhas = [wb[nome] for nome in wb.sheetnames]

    registos_validos: List[Dict[str, Any]] = []
    invalidos: Dict[str, int] = {}
    ignorados_por_empresa: Dict[str, int] = {}
    resumos_ignorados_por_empresa: Dict[str, int] = {}

    for sh in folhas:
        resultado_tabular = _processar_sheet_colunas(
            sh,
            invalidos,
            ignorados_por_empresa,
            resumos_ignorados_por_empresa,
        )

        if resultado_tabular is not None:
            registos_sheet, totais_sheet = resultado_tabular
        else:
            registos_sheet, totais_sheet = _processar_sheet_workload(
                sh,
                invalidos,
                ignorados_por_empresa,
                resumos_ignorados_por_empresa,
            )

        for empresa, tipo, canonico, minutos in registos_sheet:
            if minutos <= 0:
                continue
            empresa_norm = _norm_empresa_forte(empresa)
            if not empresa_norm:
                continue
            registos_validos.append(
                {
                    "empresa": empresa,
                    "empresa_norm": empresa_norm,
                    "tipo": tipo,
                    "canonico": canonico,
                    "minutos": minutos,
                }
            )

        for empresa, minutos in totais_sheet.items():
            if minutos <= 0:
                continue
            empresa_norm = _norm_empresa_forte(empresa)
            if not empresa_norm:
                continue
            registos_validos.append(
                {
                    "empresa": empresa,
                    "empresa_norm": empresa_norm,
                    "tipo": "resumo",
                    "canonico": None,
                    "minutos": minutos,
                }
            )

    return registos_validos, invalidos, ignorados_por_empresa, resumos_ignorados_por_empresa


def _build_timings_context(
    request: Request,
    ano_sel: Optional[int] = None,
    media_meses_sel: Optional[int] = None,
    tecnico_mapa_sel: Optional[str] = None,
    empresa_filtro: Optional[str] = None,
    tecnico_filtro: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Prepara os dados para o template timings.html / timings_mapas.html
    """
    empresa_filtro_txt = (empresa_filtro or "").strip()
    tecnico_filtro_txt = (tecnico_filtro or "").strip()
    empresa_filtro_norm = _norm_empresa_forte(empresa_filtro_txt) if empresa_filtro_txt else ""
    tecnico_filtro_canon = (
        _canonical_tecnico_nome(tecnico_filtro_txt)
        if tecnico_filtro_txt
        else ""
    )

    clientes_por_nome, clientes_norm_lista = _mapear_clientes_por_nome()
    tecnico_por_norma: Dict[str, str] = {}
    tecnicos_disponiveis_set: set[str] = set()

    for norm, cli in clientes_norm_lista:
        canonico = _tecnico_do_cliente(cli)
        tecnicos_disponiveis_set.add(canonico)
        tecnico_por_norma[norm] = canonico

    tecnicos_disponiveis_set.add(_canonical_tecnico_nome(None))

    # Anos disponíveis (só chaves numéricas)
    anos_disponiveis = (
        sorted(int(a) for a in timings_dados.keys() if str(a).isdigit())
        if timings_dados
        else []
    )

    meses_para_media = 12
    linhas = []
    ano_dict_para_mapas: dict = {}

    if not anos_disponiveis:
        ano_efetivo = ano_sel or 2025
    else:
        # escolher ano
        if ano_sel is None or ano_sel not in anos_disponiveis:
            ano_efetivo = anos_disponiveis[-1]
        else:
            ano_efetivo = ano_sel

        ano_dict_raw = timings_dados.get(str(ano_efetivo), {})
        ano_dict_resumo = dict(ano_dict_raw)
        ano_dict_para_mapas = dict(ano_dict_raw)

        # nº meses com registo (para média automática)
        meses_com_registo = set()
        for _empresa, rec in ano_dict_resumo.items():
            if rec.get("apagado"):
                continue
            meses = rec.get("meses", {})
            for num_mes, minutos in meses.items():
                try:
                    mins_int = int(minutos or 0)
                except (TypeError, ValueError):
                    mins_int = 0
                if mins_int > 0:
                    meses_com_registo.add(int(num_mes))

        if meses_com_registo:
            meses_para_media_default = len(meses_com_registo)
        else:
            meses_para_media_default = 12

        # Se o utilizador escolher na dropdown, respeitamos
        if media_meses_sel is None:
            meses_para_media = meses_para_media_default
        else:
            try:
                m = int(media_meses_sel)
            except (TypeError, ValueError):
                m = meses_para_media_default
            if m < 1 or m > 12:
                m = meses_para_media_default
            meses_para_media = m

        # construir linhas principais (por empresa) para a página de resumo
        for empresa in sorted(ano_dict_resumo.keys(), key=lambda e: e.upper()):
            rec = ano_dict_resumo[empresa]
            if rec.get("apagado"):
                continue

            empresa_norm = _norm_empresa_forte(empresa)
            tecnico_empresa = tecnico_por_norma.get(empresa_norm)
            if not tecnico_empresa:
                tecnico_empresa = _canonical_tecnico_nome(None)

            if empresa_filtro_norm and empresa_filtro_norm not in empresa_norm:
                continue

            if tecnico_filtro_canon and tecnico_empresa != tecnico_filtro_canon:
                continue

            meses_dict = rec.get("meses", {})
            extra_mensal_min = _extra_mes(rec)

            valores_por_mes = []
            minutos_por_mes = []
            total_base_min = 0

            for num_mes, _label in MESES_LABELS:
                # cuidado: depois de ler do JSON, as chaves podem vir como strings
                raw = meses_dict.get(num_mes)
                if raw is None:
                    raw = meses_dict.get(str(num_mes))
                try:
                    mins = int(raw or 0)
                except (TypeError, ValueError):
                    mins = 0

                total_base_min += mins
                mins_efetivos = mins + extra_mensal_min
                minutos_por_mes.append(mins_efetivos)
                valores_por_mes.append(_format_minutos(mins_efetivos) if mins_efetivos > 0 else "")

            total_ajustado_min = total_base_min + extra_mensal_min * 12

            if total_ajustado_min > 0 and meses_para_media > 0:
                media_min = int(round(total_ajustado_min / meses_para_media))
            else:
                media_min = 0

            linha = {
                "empresa": empresa,
                "valores_por_mes": valores_por_mes,
                "minutos_por_mes": minutos_por_mes,
                "total_base_str": _format_minutos(total_base_min)
                if total_base_min > 0
                else "",
                "extra_str": _format_minutos(extra_mensal_min) if extra_mensal_min > 0 else "",
                "total_ajustado_str": _format_minutos(total_ajustado_min)
                if total_ajustado_min > 0
                else "",
                "media_str": _format_minutos(media_min) if media_min > 0 else "",
            }
            linhas.append(linha)

    if not anos_disponiveis:
        anos_disponiveis = [ano_efetivo]
        if media_meses_sel is None:
            meses_para_media = 12

    # ========= MAPAS POR TÉCNICO / CLIENTE =========
    tecnicos_mapa: List[str] = []
    tecnico_mapa: Optional[str] = None
    mapa_cliente_tecnico: Optional[Dict[str, Any]] = None
    mapa_tecnico_global: Optional[Dict[str, Any]] = None

    if ano_dict_para_mapas:
        # mapa_raw[tecnico][empresa][mes] = minutos
        mapa_raw: dict = {}

        # 1) Empresas que têm timings (e não estão apagadas)
        for empresa, rec in ano_dict_para_mapas.items():
            if rec.get("apagado"):
                continue

            meses_dict = rec.get("meses", {})

            chave_empresa = _norm_empresa_forte(empresa)
            tecnico_key = tecnico_por_norma.get(chave_empresa)
            cliente_ref = None

            if tecnico_key is None:
                cliente_ref = clientes_por_nome.get(chave_empresa)
                if cliente_ref is None and chave_empresa:
                    for norm_cli, cli_cand in clientes_norm_lista:
                        if chave_empresa in norm_cli or norm_cli in chave_empresa:
                            cliente_ref = cli_cand
                            break
                tecnico_key = _tecnico_do_cliente(cliente_ref) if cliente_ref else _canonical_tecnico_nome(None)

            if not tecnico_key:
                tecnico_key = _canonical_tecnico_nome(None)

            for num_mes, minutos in meses_dict.items():
                try:
                    mes_int = int(num_mes)
                except (TypeError, ValueError):
                    continue
                if mes_int < 1 or mes_int > 12:
                    continue
                try:
                    mins_int = int(minutos or 0)
                except (TypeError, ValueError):
                    mins_int = 0
                if mins_int <= 0:
                    continue

                tec_dict = mapa_raw.setdefault(tecnico_key, {})
                cli_dict = tec_dict.setdefault(empresa, {})
                cli_dict[mes_int] = int(cli_dict.get(mes_int, 0) or 0) + mins_int

        # 2) Clientes com técnico mas sem timings -> queremos que apareçam no mapa,
        #    exceto se tiverem registo "apagado" no ano_dict (foram excluídos manualmente).
        nome_por_norma = {}
        for empresa_nome in ano_dict_para_mapas.keys():
            norm = _norm_empresa_forte(empresa_nome)
            if norm and norm not in nome_por_norma:
                nome_por_norma[norm] = empresa_nome

        clientes_sem_timings_por_tecnico: dict[str, list[str]] = {}
        try:
            clientes_full = estado.get("clientes", [])
        except Exception:
            clientes_full = []
        if isinstance(clientes_full, list):
            for cli in clientes_full:
                if not isinstance(cli, dict):
                    continue
                nome_cli = (
                    cli.get("nome")
                    or cli.get("cliente")
                    or cli.get("empresa")
                    or cli.get("designacao")
                )
                if not nome_cli:
                    continue

                nome_cli_str = str(nome_cli).strip()
                nome_cli_norm = _norm_empresa_forte(nome_cli_str)
                empresa_display = nome_por_norma.get(nome_cli_norm, nome_cli_str)
                rec_ano = ano_dict_para_mapas.get(empresa_display)
                # se já estiver marcado como apagado neste ano, não aparece no mapa
                if rec_ano and rec_ano.get("apagado"):
                    continue

                tecnico_key = _tecnico_do_cliente(cli)
                tec_dict_exist = mapa_raw.get(tecnico_key, {})
                ja_no_mapa = isinstance(tec_dict_exist, dict) and empresa_display in tec_dict_exist
                if ja_no_mapa:
                    continue

                lista_cli = clientes_sem_timings_por_tecnico.setdefault(tecnico_key, [])
                if empresa_display not in lista_cli:
                    lista_cli.append(empresa_display)

        # 3) Ordenar técnicos e integrar estes clientes sem timings no mapa_raw
        tecnicos_mapa = sorted({*mapa_raw.keys(), *clientes_sem_timings_por_tecnico.keys()})

        for tecnico_key, lista_clientes in clientes_sem_timings_por_tecnico.items():
            tec_dict = mapa_raw.setdefault(tecnico_key, {})
            for nome_cli_str in lista_clientes:
                tec_dict.setdefault(nome_cli_str, {})  # meses vazios (sem timings)

        if tecnicos_mapa:
            if tecnico_mapa_sel and tecnico_mapa_sel in tecnicos_mapa:
                tecnico_mapa = tecnico_mapa_sel
            else:
                tecnico_mapa = tecnicos_mapa[0]

            # -- Mapa por cliente para o técnico selecionado --
            tec_dict_sel = mapa_raw.get(tecnico_mapa, {})
            linhas_cli = []
            totais_mes_min = {num_mes: 0 for num_mes, _ in MESES_LABELS}
            total_ano_tecnico_min = 0

            for empresa in sorted(tec_dict_sel.keys(), key=lambda e: e.upper()):
                meses_cli = tec_dict_sel[empresa]
                rec_cliente = ano_dict_para_mapas.get(empresa, {})
                if rec_cliente.get("apagado"):
                    rec_cliente = {}
                extra_cli_mensal = _extra_mes(rec_cliente)

                valores_cli = []
                total_cli_base_min = 0
                tem_minutos = False
                tem_minutos_tecnico = False

                for num_mes, _label in MESES_LABELS:
                    mins_base = int(meses_cli.get(num_mes, 0) or 0)
                    if mins_base > 0:
                        tem_minutos = True
                        tem_minutos_tecnico = True
                    total_cli_base_min += mins_base

                    mins_efetivos = mins_base + extra_cli_mensal
                    if mins_efetivos > 0:
                        tem_minutos = True
                    totais_mes_min[num_mes] += mins_efetivos
                    valores_cli.append(
                        _format_minutos(mins_efetivos) if mins_efetivos > 0 else ""
                    )

                tem_extra = extra_cli_mensal > 0
                total_cli_ajustado_min = total_cli_base_min + extra_cli_mensal * 12
                total_ano_tecnico_min += total_cli_ajustado_min

                if total_cli_ajustado_min > 0:
                    media_cli_min = int(round(total_cli_ajustado_min / 12))
                else:
                    media_cli_min = 0

                sem_timings = (not tem_minutos) and (not tem_extra)
                tem_timings_outros = False
                if not tem_minutos_tecnico and isinstance(rec_cliente, dict):
                    meses_globais = rec_cliente.get("meses", {})
                    if isinstance(meses_globais, dict):
                        for valor in meses_globais.values():
                            try:
                                if int(valor or 0) > 0:
                                    tem_timings_outros = True
                                    break
                            except (TypeError, ValueError):
                                continue
                    if not tem_timings_outros and extra_cli_mensal > 0:
                        tem_timings_outros = True

                linha_cli = {
                    "cliente": empresa,
                    "valores_por_mes": valores_cli,
                    "total_str": _format_minutos(total_cli_base_min)
                    if total_cli_base_min > 0
                    else "",
                    "extra_str": _format_minutos(extra_cli_mensal)
                    if extra_cli_mensal > 0
                    else "",
                    "total_ajustado_str": _format_minutos(total_cli_ajustado_min)
                    if total_cli_ajustado_min > 0
                    else "",
                    "media_str": _format_minutos(media_cli_min) if media_cli_min > 0 else "",
                    "sem_timings": sem_timings,
                    "tem_timings_outros": tem_timings_outros,
                }
                linhas_cli.append(linha_cli)

            totais_mes_str = [
                _format_minutos(totais_mes_min[num_mes]) if totais_mes_min[num_mes] > 0 else ""
                for num_mes, _ in MESES_LABELS
            ]
            total_ano_tecnico_str = (
                _format_minutos(total_ano_tecnico_min) if total_ano_tecnico_min > 0 else ""
            )

            if total_ano_tecnico_min > 0:
                media_tec_min = int(round(total_ano_tecnico_min / 12))
            else:
                media_tec_min = 0

            mapa_cliente_tecnico = {
                "linhas": linhas_cli,
                "totais_mes_str": totais_mes_str,
                "total_ano_str": total_ano_tecnico_str,
                "media_ano_str": _format_minutos(media_tec_min) if media_tec_min > 0 else "",
            }

            # -- Mapa global por técnico --
            linhas_tec = []
            totais_mes_global_min = {num_mes: 0 for num_mes, _ in MESES_LABELS}
            total_ano_global_min = 0

            for tecnico_nome in tecnicos_mapa:
                tec_dict = mapa_raw[tecnico_nome]
                mins_por_mes = {num_mes: 0 for num_mes, _ in MESES_LABELS}
                total_tec_base_min = 0
                total_tec_extra_mensal_min = 0

                for empresa, meses_cli in tec_dict.items():
                    rec_cliente = ano_dict_para_mapas.get(empresa, {})
                    if rec_cliente.get("apagado"):
                        rec_cliente = {}
                    extra_cli_mensal = _extra_mes(rec_cliente)
                    total_tec_extra_mensal_min += extra_cli_mensal

                    for num_mes, _label in MESES_LABELS:
                        mins_base = int(meses_cli.get(num_mes, 0) or 0)
                        mins_total = mins_base + extra_cli_mensal
                        mins_por_mes[num_mes] += mins_total
                        total_tec_base_min += mins_base

                valores_mes = []
                for num_mes, _label in MESES_LABELS:
                    mins = mins_por_mes[num_mes]
                    totais_mes_global_min[num_mes] += mins
                    valores_mes.append(_format_minutos(mins) if mins > 0 else "")

                total_tec_ajustado_min = total_tec_base_min + total_tec_extra_mensal_min * 12
                total_ano_global_min += total_tec_ajustado_min

                if total_tec_ajustado_min > 0:
                    media_tec_global_min = int(round(total_tec_ajustado_min / 12))
                else:
                    media_tec_global_min = 0

                linhas_tec.append(
                    {
                        "tecnico": tecnico_nome,
                        "valores_por_mes": valores_mes,
                        "total_str": _format_minutos(total_tec_base_min)
                        if total_tec_base_min > 0
                        else "",
                        "extra_str": _format_minutos(total_tec_extra_mensal_min)
                        if total_tec_extra_mensal_min > 0
                        else "",
                        "total_ajustado_str": _format_minutos(total_tec_ajustado_min)
                        if total_tec_ajustado_min > 0
                        else "",
                        "media_str": _format_minutos(media_tec_global_min)
                        if media_tec_global_min > 0
                        else "",
                    }
                )

            totais_mes_global_str = [
                _format_minutos(totais_mes_global_min[num_mes])
                if totais_mes_global_min[num_mes] > 0
                else ""
                for num_mes, _ in MESES_LABELS
            ]
            total_ano_global_str = (
                _format_minutos(total_ano_global_min) if total_ano_global_min > 0 else ""
            )
            if total_ano_global_min > 0:
                media_global_min = int(round(total_ano_global_min / 12))
            else:
                media_global_min = 0

            mapa_tecnico_global = {
                "linhas": linhas_tec,
                "totais_mes_str": totais_mes_global_str,
                "total_ano_str": total_ano_global_str,
                "media_ano_str": _format_minutos(media_global_min) if media_global_min > 0 else "",
            }

    contexto = {
        "request": request,
        "ano_sel": ano_efetivo,
        "anos_disponiveis": anos_disponiveis,
        "meses_labels": MESES_LABELS,
        "linhas": linhas,
        "media_meses_sel": meses_para_media,
        "media_meses_opcoes": list(range(1, 13)),
        "tecnicos_mapa": tecnicos_mapa,
        "tecnico_mapa": tecnico_mapa,
        "mapa_cliente_tecnico": mapa_cliente_tecnico,
        "mapa_tecnico_global": mapa_tecnico_global,
        "empresa_q": empresa_filtro_txt,
        "tecnico_q": tecnico_filtro_canon,
        "tecnicos_opcoes": sorted(t for t in tecnicos_disponiveis_set if t),
    }
    return contexto


# ========= ROTAS =========

@router.get("/timings", response_class=HTMLResponse)
async def ver_timings(
    request: Request,
    ano: Optional[int] = None,
    media_meses: Optional[int] = None,
    tecnico_mapa: Optional[str] = None,
):
    empresa_q = (request.query_params.get("empresa_q") or "").strip()
    tecnico_q = (request.query_params.get("tecnico_q") or "").strip()

    contexto = _build_timings_context(
        request,
        ano,
        media_meses,
        tecnico_mapa,
        empresa_filtro=empresa_q,
        tecnico_filtro=tecnico_q,
    )
    return templates.TemplateResponse("timings.html", contexto)


@router.get("/timings/mapas", response_class=HTMLResponse)
async def ver_timings_mapas(
    request: Request,
    ano: Optional[int] = None,
    media_meses: Optional[int] = None,
    tecnico_mapa: Optional[str] = None,
):
    """
    Página dedicada apenas aos mapas anuais por técnico e por cliente.
    """
    contexto = _build_timings_context(request, ano, media_meses, tecnico_mapa)
    return templates.TemplateResponse("timings_mapas.html", contexto)


@router.get("/analise-grafica", response_class=HTMLResponse)
async def ver_analise_grafica(
    request: Request,
    ano: Optional[int] = None,
    media_meses: Optional[int] = None,
    tecnico_mapa: Optional[str] = None,
):
    """
    Compatibilidade com o botão 'Timings' do dashboard (antiga Análise Gráfica).
    Continua a apontar para o resumo principal.
    """
    contexto = _build_timings_context(request, ano, media_meses, tecnico_mapa)
    return templates.TemplateResponse("timings.html", contexto)


@router.post("/timings/importar")
async def importar_timings(
    request: Request,
    ano: int = Form(...),
    mes: int = Form(...),
    ficheiros: List[UploadFile] = File(...),
):
    agregados_invalidos: Dict[str, int] = {}
    agregados_ignorados_empresa: Dict[str, int] = {}
    agregados_resumos_ignorados: Dict[str, int] = {}
    duplicados_por_empresa: Dict[str, int] = {}
    armando_sem_inferido: Dict[str, int] = {}
    ficheiros_processados: List[str] = []

    registos_unicos: List[Dict[str, Any]] = []
    seen_registos: set[tuple[str, str, int, int]] = set()
    empresa_display_por_norm: Dict[str, str] = {}
    empresas_afetadas_norm: set[str] = set()
    total_minutos_deduplicados = 0
    houve_alteracoes = False

    for ficheiro in ficheiros:
        conteudo = await ficheiro.read()
        if not conteudo:
            continue

        registos, invalidos, ignorados, resumos_ignorados = _importar_excel_timings(conteudo, ano, mes)
        ficheiros_processados.append(ficheiro.filename or "sem_nome")

        for nome_norm, minutos in invalidos.items():
            agregados_invalidos[nome_norm] = agregados_invalidos.get(nome_norm, 0) + minutos
        for empresa, minutos in ignorados.items():
            agregados_ignorados_empresa[empresa] = agregados_ignorados_empresa.get(empresa, 0) + minutos
        for empresa, minutos in resumos_ignorados.items():
            agregados_resumos_ignorados[empresa] = (
                agregados_resumos_ignorados.get(empresa, 0) + minutos
            )

        for registo in registos:
            minutos = int(registo.get("minutos", 0) or 0)
            if minutos <= 0:
                continue

            empresa_norm = registo.get("empresa_norm") or ""
            if not empresa_norm:
                continue

            empresa_display = registo.get("empresa") or empresa_norm
            empresa_display_por_norm.setdefault(empresa_norm, empresa_display)

            tipo = registo.get("tipo") or "resumo"
            if tipo == "canonico":
                canonico = registo.get("canonico")
                if not canonico:
                    continue
                tecnico_chave = _norm_nome_forte(canonico)
            elif tipo == "armando":
                tecnico_chave = "__ARMANDO__"
            else:
                tecnico_chave = "__RESUMO__"

            chave_registo = (empresa_norm, tecnico_chave, mes, minutos)
            if chave_registo in seen_registos:
                empresa_repr = empresa_display_por_norm.get(empresa_norm, empresa_display)
                duplicados_por_empresa[empresa_repr] = (
                    duplicados_por_empresa.get(empresa_repr, 0) + minutos
                )
                total_minutos_deduplicados += minutos
                continue

            seen_registos.add(chave_registo)
            registos_unicos.append(registo)
            empresas_afetadas_norm.add(empresa_norm)

    if not registos_unicos and not (
        agregados_invalidos
        or agregados_ignorados_empresa
        or agregados_resumos_ignorados
        or duplicados_por_empresa
        or armando_sem_inferido
    ):
        return RedirectResponse(url=f"/timings?ano={ano}", status_code=303)

    ano_dict = _obter_ano_dict(ano)
    empresa_nome_para_inserir: Dict[str, str] = {}

    for empresa_norm in empresas_afetadas_norm:
        existente = _encontrar_empresa_existente_por_norm(ano_dict, empresa_norm)
        if existente:
            empresa_nome_para_inserir[empresa_norm] = existente
        else:
            empresa_nome_para_inserir[empresa_norm] = empresa_display_por_norm.get(
                empresa_norm, empresa_norm
            )

    for empresa_norm, empresa_nome in empresa_nome_para_inserir.items():
        rec = ano_dict.get(empresa_nome)
        if not isinstance(rec, dict):
            continue

        meses_dict = rec.get("meses")
        if isinstance(meses_dict, dict):
            meses_dict.pop(str(mes), None)
            meses_dict.pop(mes, None)
            meses_dict[int(mes)] = 0

        por_tecnico_dict = rec.get("por_tecnico")
        if isinstance(por_tecnico_dict, dict):
            for tecnico_key, tempos in por_tecnico_dict.items():
                if isinstance(tempos, dict):
                    tempos.pop(str(mes), None)
                    tempos.pop(mes, None)

    for registo in registos_unicos:
        empresa_norm = registo["empresa_norm"]
        empresa_display = registo.get("empresa") or empresa_norm
        minutos = int(registo.get("minutos", 0) or 0)
        tipo = registo.get("tipo") or "resumo"

        empresa_nome = empresa_nome_para_inserir.get(empresa_norm, empresa_display)

        tecnico_final: Optional[str]
        if tipo == "canonico":
            tecnico_final = registo.get("canonico")
        elif tipo == "armando":
            inferido = _tecnico_inferido_empresa(empresa_norm)
            if inferido:
                tecnico_final = inferido
            else:
                tecnico_final = None
                armando_sem_inferido[empresa_display] = (
                    armando_sem_inferido.get(empresa_display, 0) + minutos
                )
        else:
            tecnico_final = None

        _adicionar_tempo_empresa(ano, mes, empresa_nome, minutos, tecnico_final)
        houve_alteracoes = True

    if houve_alteracoes:
        _persistir_timings()

    if (
        agregados_invalidos
        or agregados_ignorados_empresa
        or agregados_resumos_ignorados
        or duplicados_por_empresa
        or armando_sem_inferido
    ):
        sorted_invalidos = sorted(
            agregados_invalidos.items(), key=lambda kv: kv[1], reverse=True
        )[:30]
        sorted_empresas = sorted(
            agregados_ignorados_empresa.items(), key=lambda kv: kv[1], reverse=True
        )[:30]
        sorted_resumos = sorted(
            agregados_resumos_ignorados.items(), key=lambda kv: kv[1], reverse=True
        )[:30]
        sorted_duplicados = sorted(
            duplicados_por_empresa.items(), key=lambda kv: kv[1], reverse=True
        )[:30]
        sorted_armando = sorted(
            armando_sem_inferido.items(), key=lambda kv: kv[1], reverse=True
        )[:30]

        print("[TIMINGS] Técnicos inválidos (top 30):")
        for nome, minutos in sorted_invalidos:
            print(f"  - {nome}: {minutos} min")

        print("[TIMINGS] Empresas com tempos ignorados (top 30):")
        for empresa, minutos in sorted_empresas:
            print(f"  - {empresa}: {minutos} min")

        if agregados_resumos_ignorados:
            print("[TIMINGS] Linhas-resumo ignoradas (top 30):")
            for empresa, minutos in sorted_resumos:
                print(f"  - {empresa}: {minutos} min")

        if duplicados_por_empresa:
            print("[TIMINGS] Deduplicados no lote (top 30):")
            for empresa, minutos in sorted_duplicados:
                print(f"  - {empresa}: {minutos} min")

        if armando_sem_inferido:
            print("[TIMINGS] Armando sem técnico inferido (top 30):")
            for empresa, minutos in sorted_armando:
                print(f"  - {empresa}: {minutos} min")

        relatorio = {
            "invalidos": agregados_invalidos,
            "ignorados_por_empresa": agregados_ignorados_empresa,
            "minutos_ignorados_resumo_por_empresa": agregados_resumos_ignorados,
            "duplicados_por_empresa": duplicados_por_empresa,
            "total_minutos_deduplicados": total_minutos_deduplicados,
            "armando_sem_inferido": armando_sem_inferido,
            "ficheiros": ficheiros_processados,
        }
        try:
            with open(TIMINGS_IMPORT_REPORT, "w", encoding="utf-8") as f:
                json.dump(relatorio, f, ensure_ascii=False, indent=2)
        except Exception as exc:
            print(f"[TIMINGS] Erro ao guardar relatório de importação: {exc}")

    return RedirectResponse(url=f"/timings?ano={ano}", status_code=303)


@router.post("/timings/sincronizar-clientes")
async def sincronizar_clientes(
    ano: int = Form(...),
    media_meses: Optional[int] = Form(None),
):
    """Garante que todas as empresas de estado['clientes'] existem em timings_dados."""
    try:
        clientes = estado.get("clientes", [])
    except Exception:
        clientes = []

    if not isinstance(clientes, list):
        clientes = []

    ano_int = int(ano)
    ano_dict = _obter_ano_dict(ano_int)

    normas_existentes: Dict[str, str] = {}
    for empresa_existente in list(ano_dict.keys()):
        norm = _norm_empresa_forte(empresa_existente)
        if norm and norm not in normas_existentes:
            normas_existentes[norm] = empresa_existente

    adicionados = 0

    for cli in clientes:
        if not isinstance(cli, dict):
            continue

        nome_cli = (
            cli.get("nome")
            or cli.get("cliente")
            or cli.get("empresa")
            or cli.get("designacao")
        )
        if not nome_cli:
            continue

        empresa_display = str(nome_cli).strip()
        if not empresa_display:
            continue

        norm_cli = _norm_empresa_forte(empresa_display)
        if not norm_cli:
            continue

        if norm_cli in normas_existentes:
            continue

        ano_dict[empresa_display] = {
            "meses": {num_mes: 0 for num_mes, _ in MESES_LABELS},
            "extra_mensal": 0,
            "apagado": False,
            "por_tecnico": {},
        }
        normas_existentes[norm_cli] = empresa_display
        adicionados += 1

    timings_dados[str(ano_int)] = ano_dict

    if adicionados > 0:
        print(f"[TIMINGS] Clientes sincronizados: {adicionados} novos registos")

    _persistir_timings()

    query_parts: list[str] = []
    query_parts.append(f"ano={ano_int}")
    if media_meses is not None:
        try:
            media_int = int(media_meses)
            query_parts.append(f"media_meses={media_int}")
        except (TypeError, ValueError):
            pass

    sufixo = f"?{'&'.join(query_parts)}" if query_parts else ""
    return RedirectResponse(url=f"/timings{sufixo}", status_code=303)


@router.post("/timings/migrar-formato")
async def migrar_formato_timings(
    ano: Optional[int] = Form(None),
    media_meses: Optional[int] = Form(None),
):
    """Executa migração manual e redireciona de volta à página principal."""
    sucesso, mensagem = _migrar_timings_para_minutos()
    prefixo = "concluída" if sucesso else "falhou"
    print(f"[TIMINGS] Migração manual {prefixo}: {mensagem}")

    if sucesso:
        _persistir_timings()

    query_parts: list[str] = []
    if ano is not None:
        try:
            query_parts.append(f"ano={int(ano)}")
        except (TypeError, ValueError):
            pass
    if media_meses is not None:
        try:
            query_parts.append(f"media_meses={int(media_meses)}")
        except (TypeError, ValueError):
            pass

    sufixo = f"?{'&'.join(query_parts)}" if query_parts else ""
    return RedirectResponse(url=f"/timings{sufixo}", status_code=303)

@router.post("/timings/limpar")
async def limpar_timings(
    ano: int = Form(...),
    media_meses: Optional[int] = Form(None),
    confirm_limpar: str = Form(""),
):
    """Remove todos os registos gravados e volta à página principal de timings."""
    if confirm_limpar != "APAGAR":
        raise HTTPException(status_code=400, detail="Confirmação obrigatória para limpar timings.")

    timings_dados.clear()
    _persistir_timings()

    query_parts: list[str] = []
    try:
        ano_str = str(ano).strip()
        if ano_str:
            query_parts.append(f"ano={ano_str}")
    except Exception:
        pass

    if media_meses is not None:
        query_parts.append(f"media_meses={media_meses}")

    sufixo = f"?{'&'.join(query_parts)}" if query_parts else ""
    return RedirectResponse(url=f"/timings{sufixo}", status_code=303)

@router.post("/timings/guardar")
async def guardar_timings_extras(
    request: Request,
):
    """
    Guarda os extras mensais médios por empresa (arquivo, consultoria, atendimento, etc.)
    em timings_dados.json, mantendo os minutos por mês que já estão gravados.
    """
    form = await request.form()

    ano = int(form.get("ano"))
    try:
        media_meses = int(form.get("media_meses", "12"))
    except ValueError:
        media_meses = 12

    empresas = form.getlist("empresa")
    extras_txt = form.getlist("extra")

    ano_dict = _obter_ano_dict(ano)

    for idx, nome_emp in enumerate(empresas):
        if not nome_emp:
            continue

        rec_antigo = ano_dict.get(
            nome_emp,
            {"meses": {}, "extra_mensal": 0, "apagado": False, "por_tecnico": {}},
        )

        meses_dict_antigo = rec_antigo.get("meses")
        if isinstance(meses_dict_antigo, dict):
            meses_novos = dict(meses_dict_antigo)
        else:
            meses_novos = {}

        por_tecnico_antigo = rec_antigo.get("por_tecnico")
        if isinstance(por_tecnico_antigo, dict):
            por_tecnico_novo = {
                tec: dict(meses) if isinstance(meses, dict) else {}
                for tec, meses in por_tecnico_antigo.items()
            }
        else:
            por_tecnico_novo = {}

        apagado_antigo = bool(rec_antigo.get("apagado", False))

        extra_str = extras_txt[idx] if idx < len(extras_txt) else ""
        minutos_extra = _parse_tempo_para_minutos(extra_str)

        ano_dict[nome_emp] = {
            "meses": meses_novos,
            "extra_mensal": minutos_extra,
            "apagado": apagado_antigo,
            "por_tecnico": por_tecnico_novo,
        }

    timings_dados[str(ano)] = ano_dict
    _persistir_timings()

    empresa_q = (
        request.query_params.get("empresa_q")
        or form.get("empresa_q")
        or ""
    ).strip()
    tecnico_q = (
        request.query_params.get("tecnico_q")
        or form.get("tecnico_q")
        or ""
    ).strip()

    redirect_params: Dict[str, Any] = {
        "ano": ano,
        "media_meses": media_meses,
    }
    if empresa_q:
        redirect_params["empresa_q"] = empresa_q
    if tecnico_q:
        redirect_params["tecnico_q"] = tecnico_q

    query_string = urlencode(redirect_params)
    return RedirectResponse(
        url=f"/timings?{query_string}" if query_string else "/timings",
        status_code=303,
    )


@router.post("/timings/guardar-media")
async def guardar_timings_media(
    request: Request,
):
    """
    Recebe tempos médios mensais (via campo editável em Janeiro) para clientes
    e GRAVA/ATUALIZA esses tempos em TODOS os meses do ano, mesmo que já existam
    timings anteriores. Mantém sempre o 'extra mensal' que já estava gravado e
    garante apagado=False (reativa a empresa se tinha sido apagada).
    """
    form = await request.form()
    ano = int(form.get("ano"))
    try:
        media_meses = int(form.get("media_meses", "12"))
    except ValueError:
        media_meses = 12
    tecnico_mapa = form.get("tecnico_mapa") or ""
    empresa_q = (
        request.query_params.get("empresa_q")
        or form.get("empresa_q")
        or ""
    ).strip()
    tecnico_q = (
        request.query_params.get("tecnico_q")
        or form.get("tecnico_q")
        or ""
    ).strip()

    clientes = form.getlist("cliente_media")
    jan_medias = form.getlist("jan_media")

    ano_dict = _obter_ano_dict(ano)

    for idx, nome_cli in enumerate(clientes):
        if not nome_cli:
            continue

        jan_val = jan_medias[idx] if idx < len(jan_medias) else ""
        minutos = _parse_tempo_para_minutos(jan_val)
        # Se o campo de Janeiro estiver vazio ou for inválido, não mexemos nesse cliente
        if minutos <= 0:
            continue

        rec_existente = ano_dict.get(nome_cli) or {}
        # Mantemos o extra mensal que já existia (se houver)
        extra_mensal_antigo = _extra_mes(rec_existente)

        # Sobrescrevemos TODOS os meses com o valor indicado
        meses_dict = {num_mes: minutos for num_mes, _ in MESES_LABELS}

        ano_dict[nome_cli] = {
            "meses": meses_dict,
            "extra_mensal": extra_mensal_antigo,
            "apagado": False,  # se estava apagado, volta a ficar ativo
            "por_tecnico": {},  # ao definir média manual não há detalhe por técnico
        }

    timings_dados[str(ano)] = ano_dict
    _persistir_timings()

    redirect_params: Dict[str, Any] = {
        "ano": ano,
        "media_meses": media_meses,
        "tecnico_mapa": tecnico_mapa,
    }
    if empresa_q:
        redirect_params["empresa_q"] = empresa_q
    if tecnico_q:
        redirect_params["tecnico_q"] = tecnico_q

    url = f"/timings/mapas?{urlencode(redirect_params)}"

    return RedirectResponse(url=url, status_code=303)


@router.get("/timings/excluir")
async def excluir_timing_empresa(
    request: Request,
    ano: int,
    empresa: str,
    origem: str = "timings",
    media_meses: Optional[int] = None,
    tecnico_mapa: Optional[str] = None,
):
    """
    Marca uma empresa como 'apagada' para o ano indicado (todos os meses + extra)
    e redireciona de volta para a página certa:
      - origem='timings' -> /timings
      - origem='mapas'   -> /timings/mapas

    Nota: se a empresa ainda não tiver registo nesse ano (caso venha apenas do módulo Clientes),
    cria um registo vazio com apagado=True, para não voltar a aparecer nos mapas.
    """
    ano_str = str(ano)
    empresa_q = (request.query_params.get("empresa_q") or "").strip()
    tecnico_q = (request.query_params.get("tecnico_q") or "").strip()

    ano_dict = timings_dados.get(ano_str, {})

    rec = ano_dict.get(empresa)
    if rec is None:
        rec = {"meses": {}, "extra_mensal": 0, "apagado": True}
    else:
        rec = dict(rec)
        rec["apagado"] = True
        rec["meses"] = rec.get("meses", {}) or {}
        rec["extra_mensal"] = _extra_mes(rec)

    ano_dict[empresa] = rec
    timings_dados[ano_str] = ano_dict
    _persistir_timings()

    # Normalizar media_meses
    try:
        media_val = int(media_meses) if media_meses is not None else 12
    except ValueError:
        media_val = 12

    redirect_params: Dict[str, Any] = {
        "ano": ano,
        "media_meses": media_val,
    }

    if empresa_q:
        redirect_params["empresa_q"] = empresa_q
    if tecnico_q:
        redirect_params["tecnico_q"] = tecnico_q

    if origem == "mapas":
        tec_val = (tecnico_mapa or "").strip()
        redirect_params["tecnico_mapa"] = tec_val
        url = f"/timings/mapas?{urlencode(redirect_params)}"
    else:
        url = f"/timings?{urlencode(redirect_params)}"

    return RedirectResponse(url=url, status_code=303)
